"""P2.13 — price target and return decomposition (framework 4.9).

Every expected value hand-computed. The decomposition's defining property is
that its three terms sum EXACTLY to the price change, so that is asserted at
full precision rather than to a tolerance -- a decomposition that nearly adds
up is one whose attribution nobody can trust.
"""
from decimal import Decimal

import pytest

from src.qc.rules import CLASS_A
from src.qc.valuation_rules import (
    return_decomposition_measurements,
    share_count_finding,
    target_findings,
    target_provenance_finding,
)
from src.valuation.excel.recalc import agrees, recalculate
from src.valuation.excel.workbook import (
    LAYOUT,
    SHEET_TARGET,
    ExportError,
    write_workbook,
)
from src.valuation.money import D
from src.valuation.target import (
    MULTIPLE_SOURCES,
    SOURCE_PEER_ANCHOR,
    SOURCE_REGRESSION,
    SOURCE_UNDECLARED,
    PriceTarget,
    ShareCount,
    TargetError,
    implied_price,
)


def make_target(**overrides) -> PriceTarget:
    base = dict(
        multiple=Decimal("14"), multiple_source=SOURCE_PEER_ANCHOR,
        forward_metric=Decimal("1000"), cash=Decimal("500"),
        debt=Decimal("200"),
        shares=ShareCount(Decimal("100"), Decimal("4"), Decimal("2")),
        current_price=Decimal("120"), prior_multiple=Decimal("12"),
        prior_metric=Decimal("880"), revised_metric=Decimal("940"),
        metric_label="FY26E EBITDA")
    base.update(overrides)
    return PriceTarget(**base)


class TestTheBridge:
    def test_implied_enterprise_value(self):
        assert make_target().implied_enterprise_value == Decimal("14000")

    def test_implied_equity_adds_cash_and_subtracts_debt(self):
        assert make_target().implied_equity_value == Decimal("14300")

    def test_implied_price_divides_by_fully_diluted_shares(self):
        """14300 / 106 = 134.9056..., not 14300 / 100 = 143."""
        target = make_target()
        assert target.price == Decimal("14300") / Decimal("106")
        assert target.price < Decimal("143")

    def test_upside_against_the_current_price(self):
        upside = make_target().upside
        assert (upside * 100).quantize(Decimal("0.01")) == Decimal("12.42")

    def test_a_zero_share_count_is_refused(self):
        with pytest.raises(TargetError, match="share count must be positive"):
            implied_price(D(14), D(1000), D(500), D(200), D(0))


class TestShareCount:
    def test_fully_diluted_sums_every_instrument(self):
        shares = ShareCount(Decimal("100"), Decimal("4"), Decimal("2"),
                            Decimal("10"))
        assert shares.fully_diluted == Decimal("116")

    def test_the_dilution_path_is_shown_not_asserted(self):
        rendered = ShareCount(Decimal("100"), Decimal("4"),
                              Decimal("2")).render()
        assert "100 basic" in rendered and "4 options" in rendered
        assert "6.00% dilution" in rendered

    def test_basic_only_is_a_class_a_finding(self):
        """The quiet way a retail target ends up too high."""
        finding = share_count_finding(
            make_target(shares=ShareCount(Decimal("100"))))
        assert finding is not None
        assert finding.rule.rule_class == CLASS_A
        assert "exercised into its own upside" in finding.detail

    def test_a_diluted_count_produces_no_finding(self):
        assert share_count_finding(make_target()) is None


class TestTheReturnDecomposition:
    @pytest.fixture
    def decomposition(self):
        return make_target().decomposition()

    def test_the_three_terms_sum_exactly_to_the_price_change(self, decomposition):
        """Exactly, not nearly. Asserted at full precision."""
        assert decomposition.residual == 0
        assert (decomposition.base_price + decomposition.total
                == decomposition.target_price)

    def test_each_term_is_hand_checked(self, decomposition):
        """Shares 106, cash 500, debt 200 throughout.

            base        (12 x 880 + 300) / 106 = 102.4528...
            re-rated    (14 x 880 + 300) / 106 = 119.0566...
            revised     (14 x 940 + 300) / 106 = 126.9811...
            target      (14 x 1000 + 300) / 106 = 134.9056...
        """
        assert decomposition.base_price.quantize(
            Decimal("0.0001")) == Decimal("102.4528")
        assert decomposition.re_rating.quantize(
            Decimal("0.0001")) == Decimal("16.6038")
        assert decomposition.estimate_revision.quantize(
            Decimal("0.0001")) == Decimal("7.9245")
        assert decomposition.time_roll_forward.quantize(
            Decimal("0.0001")) == Decimal("7.9245")

    def test_it_reports_in_points_of_upside(self, decomposition):
        """4.9: 'of 30% upside, 18 points is estimate revision, 12 is multiple'."""
        assert decomposition.upside_points.quantize(
            Decimal("0.1")) == Decimal("31.7")
        assert decomposition.points(decomposition.re_rating).quantize(
            Decimal("0.1")) == Decimal("16.2")

    def test_the_points_sum_to_the_total_points(self, decomposition):
        """To within an ulp, and that limit is the honest statement.

        The exact-sum invariant belongs to the currency amounts, which are
        subtractions of computed prices. Expressing each as points is a
        separate division per component, so their sum and the points of their
        sum differ in the last of fifty digits. Asserting equality here would
        be asserting something arithmetic does not provide.
        """
        parts = sum((decomposition.points(c) for c in (
            decomposition.re_rating, decomposition.estimate_revision,
            decomposition.time_roll_forward)), D(0))
        assert abs(parts - decomposition.upside_points) < Decimal("1e-45")

    def test_a_pure_re_rating_story_earns_through_the_first_term(self):
        """Maps one-to-one onto the Section 3 archetypes (4.9)."""
        rerate = make_target(prior_metric=Decimal("1000"),
                             revised_metric=Decimal("1000")).decomposition()
        assert rerate.estimate_revision == 0
        assert rerate.time_roll_forward == 0
        assert rerate.re_rating == rerate.total

    def test_a_pure_growth_story_earns_through_the_later_terms(self):
        growth = make_target(prior_multiple=Decimal("14")).decomposition()
        assert growth.re_rating == 0
        assert growth.estimate_revision + growth.time_roll_forward == growth.total

    def test_without_a_declared_revision_the_change_is_all_roll_forward(self):
        """Stated rather than assumed away."""
        rolled = make_target(revised_metric=None).decomposition()
        assert rolled.estimate_revision == 0
        assert rolled.time_roll_forward > 0

    def test_a_target_without_valuation_date_figures_refuses_to_decompose(self):
        with pytest.raises(TargetError, match="rather than an argument"):
            make_target(prior_multiple=None).decomposition()

    def test_the_measurements_report_each_component(self):
        measurements = return_decomposition_measurements(
            make_target().decomposition())
        labels = {m.label: m.value for m in measurements}
        assert labels["Upside from re-rating"] == Decimal("16.2")
        assert labels["Upside from estimate revision"] == Decimal("7.7")
        assert all(m.spec_ref == "4.9" for m in measurements)


class TestWhereTheMultipleCameFrom:
    def test_an_undeclared_source_is_a_class_a_finding(self):
        """A reverse-engineered target's arithmetic is entirely correct."""
        finding = target_provenance_finding(
            make_target(multiple_source=SOURCE_UNDECLARED))
        assert finding is not None
        assert finding.rule.rule_class == CLASS_A
        assert "only thing separating it from an honest one" in finding.detail

    @pytest.mark.parametrize("source", [s for s in MULTIPLE_SOURCES
                                        if s != SOURCE_UNDECLARED])
    def test_every_declared_source_passes(self, source):
        assert target_provenance_finding(make_target(multiple_source=source)) is None

    def test_an_invented_source_does_not_pass(self):
        finding = target_provenance_finding(
            make_target(multiple_source="felt_about_right"))
        assert finding is not None

    def test_both_rules_together(self):
        ids = {f.rule.id for f in target_findings(
            make_target(multiple_source=SOURCE_UNDECLARED,
                        shares=ShareCount(Decimal("100"))))}
        assert ids == {"share_count_not_diluted", "target_reverse_engineered"}

    def test_a_well_formed_target_raises_nothing(self):
        assert target_findings(make_target(multiple_source=SOURCE_REGRESSION)) == []


# --------------------------------------------------------------------------
# C11 coverage for this component (framework 4.9, 4.12).
# --------------------------------------------------------------------------

@pytest.fixture(scope="module")
def target_workbook(request, tmp_path_factory):
    from src.valuation.dcf import discounted_cash_flow
    from src.valuation.excel.reader import read_model
    from src.valuation.inputs import Conventions

    model = read_model(request.path.parent / "fixtures" / "tsmc_model.xlsx")
    spec = discounted_cash_flow(model.inputs, Conventions.SPEC)
    target = make_target()
    path = write_workbook(spec, tmp_path_factory.mktemp("tgt") / "m.xlsx",
                          target=target)
    return path, target


class TestExcelRecomputesTheTarget:
    @pytest.mark.parametrize("row_name,attribute", [
        ("target_price", "target_price"),
        ("target_rerating", "re_rating"),
        ("target_revision", "estimate_revision"),
        ("target_roll_forward", "time_roll_forward"),
        ("target_total", "total"),
        ("target_base", "base_price"),
    ])
    def test_each_component_reconciles(self, target_workbook, row_name, attribute):
        path, target = target_workbook
        cells = recalculate(path)
        row = getattr(LAYOUT, row_name)
        expected = getattr(target.decomposition(), attribute)
        assert agrees(expected, cells[f"{SHEET_TARGET}!B{row}".upper()])

    def test_the_upside_reconciles(self, target_workbook):
        path, target = target_workbook
        cells = recalculate(path)
        assert agrees(target.upside,
                      cells[f"{SHEET_TARGET}!B{LAYOUT.target_upside}".upper()])

    def test_excel_confirms_the_identity_on_the_face_of_the_workbook(
            self, target_workbook):
        """The last row is the three terms minus the actual price change.

        Excel computes it independently, so a decomposition that does not add
        up says so in the file rather than only in a test somebody has to run.
        """
        path, _ = target_workbook
        cells = recalculate(path)
        identity = cells[f"{SHEET_TARGET}!B{LAYOUT.target_identity}".upper()]
        assert abs(identity) < Decimal("1e-9")

    def test_excel_divides_by_fully_diluted_shares(self, target_workbook):
        path, target = target_workbook
        cells = recalculate(path)
        assert cells[f"{SHEET_TARGET}!B{LAYOUT.target_shares}".upper()] == (
            target.shares.fully_diluted)

    def test_the_target_tab_holds_no_pasted_values(self, target_workbook):
        from openpyxl import load_workbook
        path, _ = target_workbook
        sheet = load_workbook(path, data_only=False)[SHEET_TARGET]
        for row in sheet.iter_rows():
            for cell in row:
                assert not isinstance(cell.value, (int, float)), cell.coordinate

    def test_a_target_without_a_decomposition_will_not_be_exported(
            self, request, tmp_path):
        from src.valuation.dcf import discounted_cash_flow
        from src.valuation.excel.reader import read_model
        from src.valuation.inputs import Conventions

        model = read_model(request.path.parent / "fixtures" / "tsmc_model.xlsx")
        spec = discounted_cash_flow(model.inputs, Conventions.SPEC)
        with pytest.raises(ExportError, match="rather than an argument"):
            write_workbook(spec, tmp_path / "x.xlsx",
                           target=make_target(prior_multiple=None))
