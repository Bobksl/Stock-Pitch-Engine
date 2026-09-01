"""P2.11 — the reverse DCF (framework 4.7), and Excel checking its solve.

The strongest assertion here is the round trip: solve at the forward model's
own target price and the implied assumption must come back as the forecast's
own assumption. That is a genuine check on the solver rather than a check that
it returns something -- a bisection with a sign error still converges, to the
wrong number, and only the round trip notices.

The C11 coverage for this component is at the bottom. The Reverse DCF tab does
not restate Python's answer; it takes the implied growth as an input, rebuilds
the valuation, and its residual row is Excel's verdict on the solve.
"""
from dataclasses import replace
from decimal import Decimal

import pytest

from src.valuation.dcf import discounted_cash_flow
from src.valuation.excel.recalc import agrees, recalculate
from src.valuation.excel.workbook import (
    LAYOUT,
    SHEET_REVERSE,
    ExportError,
    write_workbook,
)
from src.valuation.inputs import Conventions
from src.valuation.money import D, as_percent, quantize_price
from src.valuation.reverse_dcf import (
    PRICE_TOLERANCE,
    ReverseDcf,
    ReverseDcfError,
    implied_ebitda_margin,
    implied_revenue_growth,
    implied_terminal_growth,
    reverse_dcf,
)


@pytest.fixture(scope="module")
def inputs(request):
    from src.valuation.excel.reader import read_model
    return read_model(request.path.parent / "fixtures" / "tsmc_model.xlsx").inputs


@pytest.fixture(scope="module")
def forward(inputs):
    return discounted_cash_flow(inputs, Conventions.SPEC)


class TestTheRoundTrip:
    """Solve at the model's own price; recover the model's own assumption."""

    def test_terminal_growth_round_trips(self, inputs, forward):
        solved = implied_terminal_growth(inputs, forward.share_price)
        assert abs(solved.value - inputs.terminal_growth) < Decimal("1e-8")

    def test_revenue_growth_round_trips_to_a_flat_equivalent(self, inputs, forward):
        """The fixture's growth path is not flat, so the flat equivalent sits
        inside its range rather than equalling any single year."""
        solved = implied_revenue_growth(inputs, forward.share_price)
        declared = [year.revenue_growth for year in inputs.forecast]
        assert min(declared) < solved.value < max(declared)

    def test_a_flat_forecast_round_trips_exactly(self, inputs, forward):
        """With a genuinely flat path, the solver must recover that rate."""
        flat = replace(inputs, forecast=tuple(
            replace(year, revenue_growth=Decimal("0.12"))
            for year in inputs.forecast))
        target = discounted_cash_flow(flat, Conventions.SPEC).share_price
        solved = implied_revenue_growth(flat, target)
        assert abs(solved.value - Decimal("0.12")) < Decimal("1e-6")

    def test_the_solved_price_is_inside_the_tolerance(self, inputs, forward):
        solved = implied_terminal_growth(inputs, forward.share_price)
        assert solved.residual <= PRICE_TOLERANCE


class TestMonotonicity:
    """Sanity the solver would be meaningless without."""

    def test_a_lower_price_implies_a_lower_growth_rate(self, inputs):
        cheap = implied_terminal_growth(inputs, Decimal("1100"))
        dear = implied_terminal_growth(inputs, Decimal("1900"))
        assert cheap.value < dear.value

    def test_a_lower_price_implies_a_lower_margin(self, inputs):
        cheap = implied_ebitda_margin(inputs, Decimal("1100"))
        dear = implied_ebitda_margin(inputs, Decimal("1900"))
        assert cheap.value < dear.value

    def test_a_price_far_below_the_forecast_implies_negative_growth(self, inputs):
        solved = implied_terminal_growth(inputs, Decimal("1100"))
        assert solved.value < 0


class TestTheSolverFailsRatherThanGuessing:
    def test_a_price_outside_the_bracket_is_a_named_error(self, inputs):
        """And the message says the impossibility is itself the finding."""
        with pytest.raises(ReverseDcfError, match="lies outside it"):
            implied_terminal_growth(inputs, Decimal("100000"))

    def test_the_bracket_stops_short_of_wacc(self, inputs, forward):
        """At g = WACC the perpetuity does not converge and any price is
        reachable, which would be a statement about the asymptote."""
        with pytest.raises(ReverseDcfError):
            implied_terminal_growth(
                inputs, Decimal("100000"),
                bracket=(Decimal("0"), forward.wacc - Decimal("0.01")))

    def test_an_unsolvable_assumption_is_recorded_not_dropped(self, inputs):
        """Silently omitting it would read as though it was never asked."""
        result = reverse_dcf(inputs, Decimal("100000"))
        assert "terminal_growth" in result.unsolved
        assert result.terminal_growth is None
        assert "no solution" in result.render()


class TestTheReport:
    def test_it_solves_all_three_assumptions_at_a_plausible_price(self, inputs):
        result = reverse_dcf(inputs, Decimal("1500"))
        assert len(result.assumptions) == 3
        assert result.unsolved == ()

    def test_each_assumption_is_solved_holding_the_others_fixed(self, inputs):
        """Solving jointly would have infinitely many answers and would let
        the analyst pick the flattering combination."""
        result = reverse_dcf(inputs, Decimal("1500"))
        assert as_percent(result.terminal_growth.value) == Decimal("2.35")
        assert as_percent(result.revenue_growth.value) == Decimal("12.57")
        assert as_percent(result.ebitda_margin.value) == Decimal("55.32")

    def test_the_render_names_the_price_being_explained(self, inputs):
        rendered = reverse_dcf(inputs, Decimal("1500")).render()
        assert "TWD 1500.00" in rendered
        assert "implied terminal growth" in rendered


@pytest.fixture(scope="module")
def exported(request, tmp_path_factory):
    from src.valuation.excel.reader import read_model
    model = read_model(request.path.parent / "fixtures" / "tsmc_model.xlsx")
    spec = discounted_cash_flow(model.inputs, Conventions.SPEC)
    solved = reverse_dcf(model.inputs, Decimal("1500"))
    path = write_workbook(spec, tmp_path_factory.mktemp("rev") / "m.xlsx",
                          reverse=solved)
    return path, solved


class TestExcelChecksTheSolve:
    """C11 coverage for this component (framework 4.7, 4.12)."""

    def test_excel_reproduces_the_market_price_from_the_implied_growth(
            self, exported):
        """The whole point: Excel rebuilds the valuation at Python's solved
        assumption and lands back on the price that was solved against."""
        path, solved = exported
        cells = recalculate(path)
        price = cells[f"{SHEET_REVERSE}!B{LAYOUT.reverse_share_price}".upper()]
        assert agrees(solved.market_price, price, Decimal("1e-6"))

    def test_excel_and_python_agree_on_the_residual(self, exported):
        path, solved = exported
        cells = recalculate(path)
        residual = cells[f"{SHEET_REVERSE}!B{LAYOUT.reverse_residual}".upper()]
        assert abs(abs(residual) - solved.terminal_growth.residual) < Decimal("1e-9")

    def test_the_reverse_tab_holds_no_pasted_values(self, exported):
        from openpyxl import load_workbook
        path, _ = exported
        sheet = load_workbook(path, data_only=False)[SHEET_REVERSE]
        for row in sheet.iter_rows():
            for cell in row:
                assert not isinstance(cell.value, (int, float)), cell.coordinate

    def test_an_unsolved_reverse_dcf_will_not_be_exported(self, inputs, forward,
                                                          tmp_path):
        """A tab with nothing to check would assert rather than interrogate."""
        empty = ReverseDcf(market_price=D(1500), forward=forward,
                           unsolved=("terminal_growth",))
        with pytest.raises(ExportError, match="nothing to check"):
            write_workbook(forward, tmp_path / "x.xlsx", reverse=empty)

    def test_the_workbook_is_unchanged_when_no_reverse_dcf_is_given(
            self, forward, tmp_path):
        from openpyxl import load_workbook
        path = write_workbook(forward, tmp_path / "plain.xlsx")
        assert SHEET_REVERSE not in load_workbook(path).sheetnames
