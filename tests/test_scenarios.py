"""P2.14 — scenarios and sensitivity (framework 4.10), and Excel rebuilding them.

The Sensitivity tab is the strongest C11 evidence in the repository. Changing
WACC changes every discount factor, so each of its twenty-five cells is a full
revaluation that Excel performs from first principles -- not a lookup, not a
paste. All twenty-five reconcile.
"""
from dataclasses import replace
from decimal import Decimal

import pytest

from src.qc.rules import CLASS_A
from src.qc.valuation_rules import (
    expected_value_measurement,
    scenario_findings,
    scenario_traceability_finding,
)
from src.valuation.dcf import discounted_cash_flow
from src.valuation.excel.recalc import agrees, recalculate, relative_difference
from src.valuation.excel.workbook import (
    LAYOUT,
    PENDING_SHEETS,
    SHEET_SCENARIOS,
    SHEET_SENSITIVITY,
    ExportError,
    write_workbook,
)
from src.valuation.inputs import Conventions
from src.valuation.money import D, quantize_price
from src.valuation.scenarios import (
    DEFAULT_WEIGHTS,
    PROBABILITY_STEP,
    Pillar,
    Scenario,
    ScenarioError,
    build_scenarios,
    rank_leverage,
    top_assumptions,
    two_way_table,
    validate_scenarios,
)
from src.valuation.wacc import cost_of_capital

PILLARS = (Pillar("ai_capex", "AI capex sustains leading-edge demand"),
           Pillar("node_lead", "N2 ramp keeps a one-node lead"))

SCENARIOS = (
    Scenario("bull", Decimal("0.25"), ("ai_capex", "node_lead"),
             overrides={"terminal_growth": Decimal("0.05")}),
    Scenario("base", Decimal("0.60"), ("ai_capex",)),
    Scenario("bear", Decimal("0.15"), ("node_lead",),
             overrides={"terminal_growth": Decimal("0.02")}),
)

GROWTH_AXIS = (Decimal("0.0345"), Decimal("0.0395"), Decimal("0.0445"),
               Decimal("0.0495"), Decimal("0.0545"))
WACC_AXIS = (Decimal("0.0658"), Decimal("0.0708"), Decimal("0.0759"),
             Decimal("0.0808"), Decimal("0.0858"))


@pytest.fixture(scope="module")
def inputs(request):
    from src.valuation.excel.reader import read_model
    return read_model(request.path.parent / "fixtures" / "tsmc_model.xlsx").inputs


@pytest.fixture(scope="module")
def built(inputs):
    return build_scenarios(inputs, SCENARIOS, PILLARS)


class TestScenariosTraceToPillars:
    def test_a_valid_set_validates(self, inputs):
        validate_scenarios(SCENARIOS, PILLARS)

    def test_a_case_citing_no_pillar_is_refused(self):
        orphan = (Scenario("bull", Decimal("1.0"), ()),)
        with pytest.raises(ScenarioError, match="cites no thesis pillar"):
            validate_scenarios(orphan, PILLARS)

    def test_a_case_citing_an_undeclared_pillar_is_refused(self):
        invented = (Scenario("bull", Decimal("1.0"), ("moon_landing",)),)
        with pytest.raises(ScenarioError, match="undeclared pillar"):
            validate_scenarios(invented, PILLARS)

    def test_the_rule_is_class_a(self):
        finding = scenario_traceability_finding(
            (Scenario("bull", Decimal("1.0"), ()),), PILLARS)
        assert finding is not None
        assert finding.rule.rule_class == CLASS_A
        assert "tests nothing" in finding.detail

    def test_a_traced_set_raises_nothing(self, built):
        assert scenario_findings(built) == []


class TestProbabilitiesAreCoarse:
    def test_the_default_weighting_is_the_one_the_spec_names(self):
        assert DEFAULT_WEIGHTS == {"bull": Decimal("0.25"),
                                   "base": Decimal("0.60"),
                                   "bear": Decimal("0.15")}
        assert sum(DEFAULT_WEIGHTS.values()) == 1

    def test_a_false_decimal_is_refused(self):
        """Nobody knows a thesis is 63% likely."""
        precise = (Scenario("bull", Decimal("0.63"), ("ai_capex",)),
                   Scenario("bear", Decimal("0.37"), ("ai_capex",)))
        with pytest.raises(ScenarioError, match="manufactures precision"):
            validate_scenarios(precise, PILLARS)

    def test_weights_must_sum_to_one(self):
        short = (Scenario("bull", Decimal("0.25"), ("ai_capex",)),
                 Scenario("base", Decimal("0.50"), ("ai_capex",)))
        with pytest.raises(ScenarioError, match="not an expected value"):
            validate_scenarios(short, PILLARS)

    def test_the_step_is_five_points(self):
        assert PROBABILITY_STEP == Decimal("0.05")
        for weight in DEFAULT_WEIGHTS.values():
            assert weight % PROBABILITY_STEP == 0


class TestTheWeightedOutcome:
    def test_the_bull_case_is_worth_more_than_the_bear(self, built):
        assert (built.by_name("bull").result.share_price
                > built.by_name("base").result.share_price
                > built.by_name("bear").result.share_price)

    def test_the_expected_value_is_the_weighted_sum(self, built):
        manual = sum((o.scenario.probability * o.result.share_price
                      for o in built.outcomes), D(0))
        assert built.expected_value == manual

    def test_it_sits_between_the_bear_and_the_bull(self, built):
        assert (built.by_name("bear").result.share_price
                < built.expected_value
                < built.by_name("bull").result.share_price)

    def test_the_measurement_reports_it(self, built):
        measurement = expected_value_measurement(built, "TWD")
        assert measurement.value == quantize_price(built.expected_value)
        assert measurement.spec_ref == "4.10"

    def test_a_scenario_overriding_an_unknown_field_is_refused(self, inputs):
        bad = Scenario("bull", Decimal("1.0"), ("ai_capex",),
                       overrides={"vibes": Decimal("1")})
        with pytest.raises(ScenarioError, match="no input field"):
            bad.apply(inputs)

    def test_a_driver_override_applies_to_every_year(self, inputs):
        flat = Scenario("bull", Decimal("1.0"), ("ai_capex",),
                        driver_overrides={"ebitda_margin": Decimal("0.60")})
        applied = flat.apply(inputs)
        assert {y.ebitda_margin for y in applied.forecast} == {Decimal("0.60")}

    def test_a_non_flat_driver_is_refused(self, inputs):
        bad = Scenario("bull", Decimal("1.0"), ("ai_capex",),
                       driver_overrides={"capex": Decimal("-1")})
        with pytest.raises(ScenarioError, match="not a flat driver"):
            bad.apply(inputs)


class TestLeverageRanking:
    def test_the_top_three_are_named(self, inputs):
        top = top_assumptions(inputs)
        assert len(top) == 3
        assert all(t.label for t in top)

    def test_they_are_ordered_by_how_far_the_target_moves(self, inputs):
        ranked = rank_leverage(inputs)
        moves = [r.relative_move for r in ranked]
        assert moves == sorted(moves, reverse=True)

    def test_ebitda_margin_leads_on_this_model(self, inputs):
        """A margin-driven model: the top assumption is the margin path."""
        assert top_assumptions(inputs)[0].label == "EBITDA margin"

    def test_cost_of_debt_barely_matters_at_a_three_percent_debt_weight(
            self, inputs):
        """A sanity check on the ranking: it should recover the obvious."""
        ranked = {r.label: r for r in rank_leverage(inputs)}
        assert ranked["Cost of debt"].relative_move < Decimal("0.01")
        assert ranked["Cost of debt"] is ranked[
            min(ranked, key=lambda k: ranked[k].relative_move)]

    def test_every_candidate_moves_the_price(self, inputs):
        assert all(r.move != 0 for r in rank_leverage(inputs))


@pytest.fixture(scope="module")
def table(inputs):
    return two_way_table(
        inputs, "terminal_growth", GROWTH_AXIS, "beta",
        (Decimal("1.02"), Decimal("1.22"), Decimal("1.42")),
        row_label="terminal growth", column_label="beta",
        column_is_rate=False)


class TestTwoWayTable:
    def test_it_is_the_declared_shape(self, table):
        assert len(table.prices) == 5
        assert all(len(row) == 3 for row in table.prices)

    def test_price_rises_with_growth_and_falls_with_beta(self, table):
        assert table.at(0, 0) < table.at(4, 0)          # more growth, dearer
        assert table.at(0, 0) > table.at(0, 2)          # more beta, cheaper

    def test_the_axis_format_is_declared_not_inferred(self, table):
        """A beta of 0.9 and a growth rate of 0.9 look identical to a
        heuristic and are not remotely the same quantity."""
        rendered = table.render("TWD")
        assert "1.02" in rendered and "102.00%" not in rendered
        assert "3.45%" in rendered

    def test_both_axes_must_differ(self, inputs):
        with pytest.raises(ScenarioError, match="both axes are"):
            two_way_table(inputs, "beta", (D(1),), "beta", (D(1),))


# --------------------------------------------------------------------------
# C11 coverage for this component (framework 4.10, 4.12).
# --------------------------------------------------------------------------

@pytest.fixture(scope="module")
def workbook_410(inputs, built, tmp_path_factory):
    spec = discounted_cash_flow(inputs, Conventions.SPEC)
    path = write_workbook(
        spec, tmp_path_factory.mktemp("s410") / "m.xlsx", scenarios=built,
        axes={"terminal_growth": GROWTH_AXIS, "wacc": WACC_AXIS})
    return path


@pytest.fixture(scope="module")
def recalculated_410(workbook_410):
    return recalculate(workbook_410)


class TestExcelWeightsTheScenarios:
    def test_the_weights_sum_to_one_on_the_face_of_the_sheet(
            self, recalculated_410):
        key = f"{SHEET_SCENARIOS}!B{LAYOUT.scenario_weight_check}".upper()
        assert recalculated_410[key] == 1

    def test_the_expected_value_reconciles(self, recalculated_410, built):
        key = f"{SHEET_SCENARIOS}!B{LAYOUT.scenario_expected_value}".upper()
        assert agrees(built.expected_value, recalculated_410[key])

    def test_each_contribution_reconciles(self, recalculated_410, built):
        for index, outcome in enumerate(built.outcomes):
            key = (f"{SHEET_SCENARIOS}!{LAYOUT.column(index)}"
                   f"{LAYOUT.scenario_contribution}").upper()
            assert agrees(outcome.contribution, recalculated_410[key])

    def test_the_scenarios_tab_holds_no_pasted_values(self, workbook_410):
        from openpyxl import load_workbook
        sheet = load_workbook(workbook_410, data_only=False)[SHEET_SCENARIOS]
        for row in sheet.iter_rows():
            for cell in row:
                assert not isinstance(cell.value, (int, float)), cell.coordinate


class TestExcelRebuildsTheSensitivityGrid:
    """Twenty-five full revaluations, each computed by Excel from scratch."""

    def test_every_cell_reconciles(self, recalculated_410, inputs):
        periods = len(inputs.forecast)
        first = LAYOUT.sensitivity_grid_first(periods)
        base_capital = cost_of_capital(inputs, Conventions.SPEC)

        worst = D(0)
        for row_index, growth in enumerate(GROWTH_AXIS):
            for column_index, wacc in enumerate(WACC_AXIS):
                expected = discounted_cash_flow(
                    replace(inputs, terminal_growth=growth), Conventions.SPEC,
                    capital=replace(base_capital, wacc=wacc)).share_price
                key = (f"{SHEET_SENSITIVITY}!{LAYOUT.column(column_index)}"
                       f"{first + row_index}").upper()
                worst = max(worst, relative_difference(expected,
                                                       recalculated_410[key]))
                assert agrees(expected, recalculated_410[key]), key
        # Binary64 against 50-digit Decimal across a full revaluation.
        assert worst < Decimal("1e-12")

    def test_each_column_rebuilt_its_own_discount_factors(self, recalculated_410,
                                                          inputs):
        """If the columns shared factors the grid would be a lookup table."""
        periods = len(inputs.forecast)
        row = LAYOUT.sensitivity_factor_first + periods - 1
        factors = [recalculated_410[
            f"{SHEET_SENSITIVITY}!{LAYOUT.column(i)}{row}".upper()]
            for i in range(len(WACC_AXIS))]
        assert len(set(factors)) == len(WACC_AXIS)
        assert factors == sorted(factors, reverse=True)

    def test_the_sensitivity_tab_holds_no_pasted_values(self, workbook_410):
        from openpyxl import load_workbook
        sheet = load_workbook(workbook_410, data_only=False)[SHEET_SENSITIVITY]
        for row in sheet.iter_rows():
            for cell in row:
                assert not isinstance(cell.value, (int, float)), cell.coordinate

    def test_one_axis_is_not_a_two_way_table(self, inputs, tmp_path):
        spec = discounted_cash_flow(inputs, Conventions.SPEC)
        with pytest.raises(ExportError, match="one axis is not two-way"):
            write_workbook(spec, tmp_path / "x.xlsx",
                           axes={"terminal_growth": GROWTH_AXIS})


class TestEveryTabTheFrameworkNamesExists:
    def test_nothing_is_still_pending(self):
        """4.12 names eight tabs. This asserts the list is finally empty."""
        assert PENDING_SHEETS == ()

    def test_the_workbook_carries_them_all(self, workbook_410, inputs,
                                           tmp_path, built):
        from openpyxl import load_workbook

        from src.valuation.comps import (
            METRIC_SALES, NUMERATOR_ENTERPRISE, CompsExport, Peer)
        from src.valuation.reverse_dcf import reverse_dcf
        from src.valuation.target import (
            SOURCE_PEER_ANCHOR, PriceTarget, ShareCount)

        spec = discounted_cash_flow(inputs, Conventions.SPEC)
        peers = (Peer("A", Decimal("5.0"), D(1), Decimal("0.19"), None,
                      "CY2024", "consensus"),
                 Peer("B", Decimal("3.6"), D(1), Decimal("0.16"), None,
                      "CY2024", "consensus"))
        everything = write_workbook(
            spec, tmp_path / "all.xlsx",
            reverse=reverse_dcf(inputs, Decimal("1500")),
            comps=CompsExport(peers, NUMERATOR_ENTERPRISE, METRIC_SALES,
                              Decimal("1000"), Decimal("0.25")),
            target=PriceTarget(
                multiple=Decimal("14"), multiple_source=SOURCE_PEER_ANCHOR,
                forward_metric=Decimal("1000"), cash=Decimal("500"),
                debt=Decimal("200"), shares=ShareCount(D(100), D(6)),
                current_price=Decimal("120"), prior_multiple=Decimal("12"),
                prior_metric=Decimal("880")),
            scenarios=built,
            axes={"terminal_growth": GROWTH_AXIS, "wacc": WACC_AXIS})

        assert set(load_workbook(everything).sheetnames) == {
            "Inputs", "WACC", "Model", "DCF", "Comps", "Target", "Scenarios",
            "Sensitivity", "Reverse DCF", "Summary"}
