"""P2.12 — comparable companies (framework 4.8), and Excel recomputing them.

The Reddit-at-IPO arithmetic is encoded here as constants. The source is a
newsletter and is deliberately not committed; what matters is reproducible
from the numbers the framework itself quotes, and those are hand-checked:

    growth-adjusted   PINS 5.0/19 = 0.263x   SNAP 3.6/16 = 0.225x
    anchored to PINS  0.263 x 25 = 6.58x  →  $41.43-44.10, 22-30% upside
    anchored to SNAP  0.225 x 25 = 5.62x  →  $36.76,        +8% upside

Same method, same inputs, same day. The headline is roughly three times larger
purely because of which peer was chosen, and the source never flags it. That
is the whole argument for mandatory anchor disclosure, so it is a test.
"""
from decimal import Decimal

import pytest

from src.qc.rules import CLASS_A, CLASS_B
from src.qc.valuation_rules import (
    anchor_disclosure_finding,
    comp_consistency_finding,
    comp_set_size_finding,
    comps_findings,
    pairing_finding,
)
from src.valuation.comps import (
    METRIC_BOOK,
    METRIC_EARNINGS,
    METRIC_EBIT,
    METRIC_EBITDA,
    METRIC_FCFE,
    METRIC_SALES,
    MINIMUM_COMP_SET,
    NUMERATOR_ENTERPRISE,
    NUMERATOR_EQUITY,
    CompSetError,
    CompsExport,
    PairingError,
    Peer,
    TargetProfile,
    anchor_disclosure,
    check_pairing,
    label,
    outside_peer_distribution,
    regress,
    select_metric,
)
from src.valuation.money import D

# Framework 4.8's worked example, as constants. No PDF is committed.
PINS = Peer("PINS", Decimal("5.0"), D(1), Decimal("0.19"), Decimal("0.20"),
            "CY2024", "consensus")
SNAP = Peer("SNAP", Decimal("3.6"), D(1), Decimal("0.16"), Decimal("0.10"),
            "CY2024", "consensus")
RDDT_GROWTH = Decimal("0.25")

#: A wide-enough set for a two-regressor fit.
WIDE = (
    Peer("A", Decimal("50"), D(10), Decimal("0.30"), Decimal("0.35"),
         "CY2024", "consensus"),
    Peer("B", Decimal("36"), D(10), Decimal("0.22"), Decimal("0.28"),
         "CY2024", "consensus"),
    Peer("C", Decimal("24"), D(10), Decimal("0.15"), Decimal("0.20"),
         "CY2024", "consensus"),
    Peer("D", Decimal("60"), D(10), Decimal("0.35"), Decimal("0.42"),
         "CY2024", "consensus"),
    Peer("E", Decimal("18"), D(10), Decimal("0.10"), Decimal("0.15"),
         "CY2024", "consensus"),
    Peer("F", Decimal("41"), D(10), Decimal("0.26"), Decimal("0.30"),
         "CY2024", "consensus"),
)


class TestThePairingRuleIsAHardError:
    @pytest.mark.parametrize("metric", [METRIC_SALES, METRIC_EBITDA, METRIC_EBIT])
    def test_enterprise_value_pairs_with_pre_interest_metrics(self, metric):
        check_pairing(NUMERATOR_ENTERPRISE, metric)

    @pytest.mark.parametrize("metric", [METRIC_EARNINGS, METRIC_FCFE, METRIC_BOOK])
    def test_equity_value_pairs_with_post_interest_metrics(self, metric):
        check_pairing(NUMERATOR_EQUITY, metric)

    def test_ev_over_earnings_raises_rather_than_warning(self):
        """Not conservative, not aggressive -- a ratio across the structure."""
        with pytest.raises(PairingError) as exc:
            check_pairing(NUMERATOR_ENTERPRISE, METRIC_EARNINGS)
        assert "equity value metric" in str(exc.value)

    def test_equity_over_ebitda_raises(self):
        with pytest.raises(PairingError, match="enterprise value metric"):
            check_pairing(NUMERATOR_EQUITY, METRIC_EBITDA)

    def test_the_registry_rule_is_still_reachable(self):
        """It raises first; the finding exists so the rule is maintainable."""
        finding = pairing_finding(NUMERATOR_ENTERPRISE, METRIC_EARNINGS)
        assert finding is not None and finding.rule.rule_class == CLASS_A

    def test_a_legal_pairing_produces_no_finding(self):
        assert pairing_finding(NUMERATOR_ENTERPRISE, METRIC_EBITDA) is None


class TestTheMetricDecisionTree:
    @pytest.mark.parametrize("profile,expected", [
        (TargetProfile(profitable=True, depreciation_material=True),
         (NUMERATOR_ENTERPRISE, METRIC_EBITDA)),
        (TargetProfile(profitable=True, depreciation_material=False),
         (NUMERATOR_ENTERPRISE, METRIC_EBIT)),
        (TargetProfile(profitable=False, depreciation_material=True,
                       high_growth_long_runway=True),
         (NUMERATOR_ENTERPRISE, METRIC_SALES)),
        (TargetProfile(profitable=True, depreciation_material=True,
                       earnings_distorted_by_reinvestment=True),
         (NUMERATOR_EQUITY, METRIC_FCFE)),
        (TargetProfile(profitable=True, depreciation_material=True,
                       is_financial=True),
         (NUMERATOR_EQUITY, METRIC_EARNINGS)),
    ])
    def test_each_row_of_the_table(self, profile, expected):
        assert select_metric(profile) == expected

    def test_financials_override_every_other_row(self):
        """Interest is operating, not financing: EV multiples are meaningless."""
        profile = TargetProfile(profitable=False, depreciation_material=False,
                                high_growth_long_runway=True, is_financial=True)
        assert select_metric(profile)[0] == NUMERATOR_EQUITY

    def test_a_profile_the_table_does_not_cover_refuses_to_guess(self):
        with pytest.raises(CompSetError, match="prescribes no multiple"):
            select_metric(TargetProfile(profitable=False,
                                        depreciation_material=True))

    def test_every_selected_pair_satisfies_the_pairing_rule(self):
        for profile in (TargetProfile(True, True), TargetProfile(True, False),
                        TargetProfile(False, True, high_growth_long_runway=True),
                        TargetProfile(True, True,
                                      earnings_distorted_by_reinvestment=True),
                        TargetProfile(True, True, is_financial=True)):
            check_pairing(*select_metric(profile))


class TestTheRedditWorkedExample:
    def test_the_growth_adjusted_multiples(self):
        assert PINS.growth_adjusted().quantize(Decimal("0.001")) == Decimal("0.263")
        assert SNAP.growth_adjusted().quantize(Decimal("0.001")) == Decimal("0.225")

    def test_reddits_own_growth_adjusted_multiple(self):
        """5.11 / 25 = 0.204x, the third figure the framework quotes."""
        assert (Decimal("5.11") / Decimal(25)).quantize(
            Decimal("0.001")) == Decimal("0.204")

    def test_the_implied_multiple_under_each_anchor(self):
        disclosure = anchor_disclosure(
            (PINS, SNAP), RDDT_GROWTH, D(1), NUMERATOR_ENTERPRISE, METRIC_SALES)
        implied = {a.peer: a.implied_multiple.quantize(Decimal("0.01"))
                   for a in disclosure.anchors}
        assert implied == {"PINS": Decimal("6.58"), "SNAP": Decimal("5.62")}

    def test_the_headline_upside_is_roughly_three_times_larger(self):
        """22-30% against +8%, on the same inputs. This is the finding."""
        current = Decimal("33.95")
        pins_high = (Decimal("44.10") / current - 1) * 100
        snap = (Decimal("36.76") / current - 1) * 100
        assert pins_high.quantize(Decimal("0.1")) == Decimal("29.9")
        assert snap.quantize(Decimal("0.1")) == Decimal("8.3")
        assert pins_high / snap > 3


class TestAnchorDisclosure:
    @pytest.fixture
    def disclosure(self):
        return anchor_disclosure(
            (PINS, SNAP), RDDT_GROWTH, Decimal("1000"), NUMERATOR_ENTERPRISE,
            METRIC_SALES, net_debt=Decimal("-500"), shares=Decimal("100"),
            current_price=Decimal("50"))

    def test_every_peer_appears(self, disclosure):
        assert {a.peer for a in disclosure.anchors} == {"PINS", "SNAP"}

    def test_the_range_and_the_spread_are_reported(self, disclosure):
        assert disclosure.low.peer == "SNAP"
        assert disclosure.high.peer == "PINS"
        assert disclosure.spread.quantize(Decimal("0.01")) == Decimal("1.17")

    def test_net_debt_is_added_back_on_the_enterprise_side(self, disclosure):
        """Signed: net debt of -500 is a net cash position."""
        pins = next(a for a in disclosure.anchors if a.peer == "PINS")
        assert pins.implied_price == (pins.implied_numerator + 500) / 100

    def test_the_render_shows_the_range_not_just_one_anchor(self, disclosure):
        rendered = disclosure.render("USD")
        assert "PINS" in rendered and "SNAP" in rendered
        assert "a factor of 1.17 on the choice of peer" in rendered

    def test_reporting_one_anchor_is_a_class_a_finding(self, disclosure):
        finding = anchor_disclosure_finding(disclosure, ("PINS",))
        assert finding is not None
        assert finding.rule.rule_class == CLASS_A
        assert "SNAP omitted" in finding.detail
        assert "factor of 1.17" in finding.detail

    def test_reporting_every_anchor_passes(self, disclosure):
        assert anchor_disclosure_finding(disclosure, ("PINS", "SNAP")) is None

    def test_reporting_none_is_caught(self, disclosure):
        assert anchor_disclosure_finding(disclosure, None) is not None


class TestTheRegressionIsThePrimaryMethod:
    def test_it_fits_growth_and_profitability(self):
        fitted = regress(WIDE)
        assert fitted.profitability_coefficient is not None
        assert fitted.growth_coefficient > 0

    def test_residuals_sum_to_approximately_zero(self):
        """A property of least squares with an intercept; a sign error breaks it."""
        fitted = regress(WIDE)
        total = sum(fitted.residuals.values(), D(0))
        assert abs(total) < Decimal("1e-20")

    def test_a_peer_on_the_line_has_no_residual(self):
        """Construct a set lying exactly on a plane; every residual is zero."""
        exact = tuple(
            Peer(f"P{i}", Decimal(2) + Decimal(10) * Decimal(f"0.{i}0"), D(1),
                 Decimal(f"0.{i}0"), None, "CY2024", "consensus")
            for i in range(1, 6))
        fitted = regress(exact, use_profitability=False)
        assert all(abs(r) < Decimal("1e-20") for r in fitted.residuals.values())

    def test_the_residual_identifies_cheap_and_dear(self, ):
        fitted = regress(WIDE)
        cheapest = min(fitted.residuals, key=fitted.residuals.get)
        assert fitted.residuals[cheapest] < 0
        assert cheapest in fitted.render()

    def test_too_few_peers_to_determine_the_parameters_is_refused(self):
        """Fitting a line through as many points as it has parameters
        describes the points, not the relationship."""
        with pytest.raises(CompSetError, match="degrees of freedom"):
            regress(WIDE[:3])

    def test_peers_that_do_not_vary_independently_are_refused(self):
        singular = tuple(
            Peer(f"S{i}", Decimal("30"), D(10), Decimal("0.20"), Decimal("0.25"),
                 "CY2024", "consensus") for i in range(6))
        with pytest.raises(CompSetError, match="singular"):
            regress(singular)

    def test_profitability_is_dropped_when_a_peer_lacks_it(self):
        partial = WIDE[:-1] + (
            Peer("G", Decimal("30"), D(10), Decimal("0.20"), None,
                 "CY2024", "consensus"),)
        assert regress(partial).profitability_coefficient is None


class TestCompSetConsistency:
    def test_mixed_periods_are_a_class_a_finding(self):
        mixed = (PINS, Peer("SNAP", Decimal("3.6"), D(1), Decimal("0.16"),
                            None, "CY2023", "consensus"))
        finding = comp_consistency_finding(mixed)
        assert finding is not None and finding.rule.rule_class == CLASS_A
        assert "calendar periods" in finding.detail

    def test_mixed_estimate_sources_are_caught(self):
        mixed = (PINS, Peer("SNAP", Decimal("3.6"), D(1), Decimal("0.16"),
                            None, "CY2024", "BEst"))
        assert "estimate sources" in comp_consistency_finding(mixed).detail

    def test_a_consistent_set_produces_no_finding(self):
        assert comp_consistency_finding((PINS, SNAP)) is None

    def test_anchor_disclosure_refuses_an_inconsistent_set_outright(self):
        mixed = (PINS, Peer("SNAP", Decimal("3.6"), D(1), Decimal("0.16"),
                            None, "CY2023", "consensus"))
        with pytest.raises(CompSetError, match="different calendar periods"):
            anchor_disclosure(mixed, RDDT_GROWTH, D(1), NUMERATOR_ENTERPRISE,
                              METRIC_SALES)


class TestCompSetSize:
    def test_below_the_minimum_is_class_b_and_exceptionable(self):
        finding = comp_set_size_finding((PINS, SNAP))
        assert finding is not None
        assert finding.rule.rule_class == CLASS_B
        assert finding.rule.exceptionable
        assert "teaching example" in finding.detail

    def test_at_the_minimum_it_passes(self):
        assert comp_set_size_finding(WIDE[:MINIMUM_COMP_SET]) is None


class TestTheSanityGate:
    def test_a_multiple_inside_the_peer_range_passes(self):
        assert outside_peer_distribution(Decimal("4.0"), (PINS, SNAP)) is False

    def test_a_multiple_above_every_peer_is_flagged(self):
        assert outside_peer_distribution(Decimal("9.0"), (PINS, SNAP)) is True

    def test_a_multiple_below_every_peer_is_flagged(self):
        assert outside_peer_distribution(Decimal("1.0"), (PINS, SNAP)) is True


class TestAllRulesTogether:
    def test_an_inconsistent_thin_set_raises_both_findings(self):
        comps = CompsExport(
            peers=(PINS, Peer("SNAP", Decimal("3.6"), D(1), Decimal("0.16"),
                              None, "CY2023", "BEst")),
            numerator=NUMERATOR_ENTERPRISE, metric=METRIC_SALES,
            target_metric_value=Decimal("1000"), target_growth=RDDT_GROWTH)
        ids = {f.rule.id for f in comps_findings(comps)}
        assert ids == {"comp_definitions_inconsistent", "comp_set_below_minimum"}

    def test_a_clean_wide_set_with_full_disclosure_raises_nothing(self):
        comps = CompsExport(
            peers=WIDE, numerator=NUMERATOR_ENTERPRISE, metric=METRIC_SALES,
            target_metric_value=Decimal("1000"), target_growth=RDDT_GROWTH)
        reported = tuple(p.ticker for p in WIDE)
        assert comps_findings(comps, reported_anchors=reported) == []

    def test_labels_are_the_conventional_names(self):
        assert label(NUMERATOR_ENTERPRISE, METRIC_EBITDA) == "EV/EBITDA"
        assert label(NUMERATOR_EQUITY, METRIC_FCFE) == "P/FCF"


# --------------------------------------------------------------------------
# C11 coverage for this component (framework 4.8, 4.12).
#
# The tab computes every anchor from the peer table as live formulas, and the
# range and spread on the bottom rows come from MIN/MAX over the row above --
# so a reader cannot take one column home without the others being on screen.
# --------------------------------------------------------------------------

@pytest.fixture(scope="module")
def comps_export():
    return CompsExport(
        peers=(PINS, SNAP), numerator=NUMERATOR_ENTERPRISE, metric=METRIC_SALES,
        target_metric_value=Decimal("1000"), target_growth=RDDT_GROWTH,
        net_debt=Decimal("-1620"), current_price=Decimal("100"))


@pytest.fixture(scope="module")
def comps_workbook(request, tmp_path_factory, comps_export):
    from src.valuation.dcf import discounted_cash_flow
    from src.valuation.excel.reader import read_model
    from src.valuation.excel.workbook import write_workbook
    from src.valuation.inputs import Conventions

    model = read_model(request.path.parent / "fixtures" / "tsmc_model.xlsx")
    spec = discounted_cash_flow(model.inputs, Conventions.SPEC)
    path = write_workbook(spec, tmp_path_factory.mktemp("comps") / "m.xlsx",
                          comps=comps_export)
    return path, model.inputs.shares_outstanding


class TestExcelRecomputesTheAnchors:
    def test_every_anchor_reconciles(self, comps_workbook, comps_export):
        from src.valuation.excel.recalc import agrees, recalculate
        from src.valuation.excel.workbook import LAYOUT, SHEET_COMPS

        path, shares = comps_workbook
        cells = recalculate(path)
        disclosure = comps_export.disclosure(shares=shares)

        for index, anchor in enumerate(disclosure.anchors):
            column = LAYOUT.column(index)
            for row, expected in (
                    (LAYOUT.comps_multiple, anchor.peer_multiple),
                    (LAYOUT.comps_growth_adjusted, anchor.peer_growth_adjusted),
                    (LAYOUT.comps_implied_multiple, anchor.implied_multiple),
                    (LAYOUT.comps_implied_price, anchor.implied_price),
                    (LAYOUT.comps_upside, anchor.upside)):
                key = f"{SHEET_COMPS}!{column}{row}".upper()
                assert agrees(expected, cells[key]), f"{key} for {anchor.peer}"

    def test_the_range_and_spread_reconcile(self, comps_workbook, comps_export):
        from src.valuation.excel.recalc import agrees, recalculate
        from src.valuation.excel.workbook import LAYOUT, SHEET_COMPS

        path, shares = comps_workbook
        cells = recalculate(path)
        disclosure = comps_export.disclosure(shares=shares)

        assert agrees(disclosure.low.implied_multiple,
                      cells[f"{SHEET_COMPS}!B{LAYOUT.comps_range_low}".upper()])
        assert agrees(disclosure.high.implied_multiple,
                      cells[f"{SHEET_COMPS}!B{LAYOUT.comps_range_high}".upper()])
        assert agrees(disclosure.spread,
                      cells[f"{SHEET_COMPS}!B{LAYOUT.comps_spread}".upper()])

    def test_the_comps_tab_holds_no_pasted_values(self, comps_workbook):
        from openpyxl import load_workbook
        from src.valuation.excel.workbook import SHEET_COMPS

        path, _ = comps_workbook
        sheet = load_workbook(path, data_only=False)[SHEET_COMPS]
        for row in sheet.iter_rows():
            for cell in row:
                assert not isinstance(cell.value, (int, float)), cell.coordinate

    def test_a_single_peer_will_not_be_exported(self, request, tmp_path):
        """One anchor has no range, and the range is the deliverable."""
        from src.valuation.dcf import discounted_cash_flow
        from src.valuation.excel.reader import read_model
        from src.valuation.excel.workbook import ExportError, write_workbook
        from src.valuation.inputs import Conventions

        model = read_model(request.path.parent / "fixtures" / "tsmc_model.xlsx")
        spec = discounted_cash_flow(model.inputs, Conventions.SPEC)
        lonely = CompsExport(peers=(PINS,), numerator=NUMERATOR_ENTERPRISE,
                             metric=METRIC_SALES,
                             target_metric_value=Decimal("1000"),
                             target_growth=RDDT_GROWTH)
        with pytest.raises(ExportError, match="at least two peers"):
            write_workbook(spec, tmp_path / "one.xlsx", comps=lonely)
