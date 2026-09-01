"""P2.5 — C11: Excel recalculation matches Python within tolerance (4.12).

The test that makes the dual-track design safe. Python owns the calculation
and the workbook carries live formulas, so the two can drift; C11 is what
notices. It runs at two levels deliberately.

**Per operation.** Each entry in the paired vocabulary is recalculated on its
own in a two-cell workbook. When a pairing breaks, this names the operation.

**End to end.** The whole exported model is recalculated and every headline
compared against Python computed from the workbook's own inputs -- read back
through `readback.py`, not the inputs the export started from, so the
comparison isolates formula divergence from input quantization.

A missing engine fails these tests rather than skipping them. A skipped
reconciliation is a gate that has stopped running while still reporting green.
"""
from decimal import Decimal

import pytest

from src.valuation.dcf import discounted_cash_flow
from src.valuation.excel.formulas import OPS, FormulaError, compute, formula
from src.valuation.excel.readback import ReadbackError, read_inputs
from src.valuation.excel.recalc import (
    C11_TOLERANCE,
    ENGINE_NONE,
    RecalculationUnavailable,
    agrees,
    available_engines,
    recalculate,
    relative_difference,
    selected_engine,
)
from src.valuation.excel.workbook import (
    LAYOUT,
    SHEET_DCF,
    SHEET_SUMMARY,
    SHEET_WACC,
    ExportError,
    write_workbook,
)
from src.valuation.inputs import Conventions
from src.valuation.money import to_spreadsheet

#: Sample operands per operation, chosen to be ordinary rather than adversarial:
#: the point is to catch a transcription error in a formula, not to probe
#: floating-point edges the workbook will never see.
SAMPLES = {
    "sum": (Decimal("1200.5"), Decimal("-300.25"), Decimal("42")),
    "difference": (Decimal("2100.75"), Decimal("880.5")),
    "product": (Decimal("3959.41"), Decimal("0.6733")),
    "ratio": (Decimal("51666.26"), Decimal("59557.64")),
    "growth": (Decimal("4533.53"), Decimal("3959.41")),
    "grow_by": (Decimal("2894.3077"), Decimal("0.368")),
    "complement": (Decimal("0.0278"),),
    "negated_product": (Decimal("1873.97"), Decimal("0.166")),
    "weighted_pair": (Decimal("0.076996"), Decimal("0.9722"),
                      Decimal("0.034961"), Decimal("0.0278")),
    "discount_factor": (Decimal("0.0758268"), Decimal("5.1666666666666667")),
    "perpetuity": (Decimal("2260.5609"), Decimal("0.0445"), Decimal("0.0758268")),
    # Comps operations (4.8). per_point and times_points are inverses, so the
    # samples are the Reddit worked example in both directions.
    "minimum": (Decimal("6.58"), Decimal("5.62"), Decimal("4.10")),
    "maximum": (Decimal("6.58"), Decimal("5.62"), Decimal("4.10")),
    "per_point": (Decimal("5.0"), Decimal("0.19")),
    "times_points": (Decimal("0.2631578947368421"), Decimal("0.25")),
    "upside": (Decimal("316.195424929466"), Decimal("100")),
    # The 4.9 price bridge: multiple, metric, cash, debt, fully diluted shares.
    "implied_price": (Decimal("14"), Decimal("1000"), Decimal("500"),
                      Decimal("200"), Decimal("106")),
}


@pytest.fixture(scope="module")
def spec_result(request):
    from src.valuation.excel.reader import read_model
    workbook = request.path.parent / "fixtures" / "tsmc_model.xlsx"
    model = read_model(workbook)
    return model, discounted_cash_flow(model.inputs, Conventions.SPEC)


@pytest.fixture(scope="module")
def exported(spec_result, tmp_path_factory):
    _, result = spec_result
    return write_workbook(result, tmp_path_factory.mktemp("c11") / "model.xlsx")


@pytest.fixture(scope="module")
def recalculated(exported):
    return recalculate(exported)


class TestTheEngineIsRequired:
    def test_an_engine_is_available(self):
        """If this fails, C11 cannot run and the suite must say so loudly."""
        assert available_engines(), (
            "no recalculation engine: pip install formulas, or install "
            "LibreOffice. C11 is not skippable.")

    def test_selecting_an_unavailable_engine_is_an_error(self, monkeypatch):
        monkeypatch.setenv("QC_RECALC_ENGINE", "abacus")
        with pytest.raises(RecalculationUnavailable, match="not available"):
            selected_engine()

    def test_disabling_is_explicit_and_returns_a_distinguishable_result(
            self, monkeypatch, exported):
        """Opting out is possible; it must not look like a clean pass."""
        monkeypatch.setenv("QC_RECALC_ENGINE", ENGINE_NONE)
        assert selected_engine() == ENGINE_NONE
        assert recalculate(exported) is None


class TestEachOperationPairs:
    @pytest.mark.parametrize("name", sorted(SAMPLES))
    def test_python_and_excel_agree_on_one_operation(self, name, tmp_path):
        """A pairing bug names the operation rather than the workbook."""
        from openpyxl import Workbook

        values = SAMPLES[name]
        book = Workbook()
        sheet = book.active
        sheet.title = "OP"
        for index, value in enumerate(values, start=1):
            sheet.cell(index, 1, to_spreadsheet(value))
        refs = [f"A{index}" for index in range(1, len(values) + 1)]
        sheet.cell(1, 3, formula(name, *refs))

        path = tmp_path / f"{name}.xlsx"
        book.save(path)

        expected = compute(name, *(to_spreadsheet(v) for v in values))
        actual = recalculate(path)["OP!C1"]
        assert agrees(expected, actual), (
            f"{name}: python {expected} vs excel {actual}, "
            f"relative {relative_difference(expected, actual)}")

    def test_every_operation_has_a_sample(self):
        """A new operation without a pairing test would go unchecked."""
        assert set(SAMPLES) == set(OPS)

    def test_wrong_arity_is_refused(self):
        with pytest.raises(FormulaError, match="takes 2 arguments"):
            formula("ratio", "A1")

    def test_an_unknown_operation_names_the_file_to_change(self):
        with pytest.raises(FormulaError, match="both implementations"):
            formula("black_scholes", "A1", "A2")


@pytest.fixture(scope="module")
def python_side(exported, spec_result):
    """Python computed from the workbook's OWN inputs, read back.

    Comparing against the pre-export result instead would fold input
    quantization into the same number as formula divergence.
    """
    model, _ = spec_result
    return discounted_cash_flow(read_inputs(exported, model.inputs),
                                Conventions.SPEC)


class TestEndToEnd:
    @pytest.mark.parametrize("cell,attribute", [
        (f"{SHEET_WACC}!B{LAYOUT.wacc_value}", "wacc"),
        (f"{SHEET_DCF}!B{LAYOUT.dcf_pv_forecast}", "pv_forecast"),
        (f"{SHEET_DCF}!B{LAYOUT.dcf_terminal_value}", "terminal_value"),
        (f"{SHEET_DCF}!B{LAYOUT.dcf_pv_terminal}", "pv_terminal_value"),
        (f"{SHEET_DCF}!B{LAYOUT.dcf_enterprise_value}", "enterprise_value"),
        (f"{SHEET_DCF}!B{LAYOUT.dcf_equity_value}", "equity_value"),
        (f"{SHEET_DCF}!B{LAYOUT.dcf_share_price}", "share_price"),
        (f"{SHEET_SUMMARY}!B4", "share_price"),
    ])
    def test_headline_cells_reconcile(self, recalculated, python_side,
                                      cell, attribute):
        expected = getattr(python_side, attribute)
        actual = recalculated[cell.upper()]
        assert agrees(expected, actual), (
            f"{cell}: python {expected} vs excel {actual}, "
            f"relative {relative_difference(expected, actual)} "
            f"exceeds {C11_TOLERANCE}")

    def test_every_forecast_period_reconciles(self, recalculated, python_side):
        from src.valuation.excel.workbook import SHEET_MODEL
        for index, period in enumerate(python_side.periods):
            column = LAYOUT.column(index)
            for row, value in ((LAYOUT.model_revenue, period.revenue),
                               (LAYOUT.model_ebitda, period.ebitda),
                               (LAYOUT.model_ebit, period.ebit),
                               (LAYOUT.model_ufcf, period.unlevered_fcf)):
                key = f"{SHEET_MODEL}!{column}{row}".upper()
                assert agrees(value, recalculated[key]), (
                    f"{key} for {period.period}: python {value} vs "
                    f"excel {recalculated[key]}")

    def test_the_agreement_is_far_tighter_than_the_tolerance(
            self, recalculated, python_side):
        """Binary64 against 50-digit Decimal: the real gap is ~1e-16."""
        key = f"{SHEET_DCF}!B{LAYOUT.dcf_share_price}".upper()
        difference = relative_difference(python_side.share_price,
                                         recalculated[key])
        assert difference < Decimal("1e-12")

    def test_the_target_price_is_the_expected_one(self, python_side):
        from src.valuation.money import quantize_price
        assert quantize_price(python_side.share_price) == Decimal("2321.64")


class TestCalculationTabsHoldFormulasNotValues:
    def test_no_calculation_cell_is_a_pasted_number(self, exported):
        """The property that makes C11 mean anything.

        Were Python's results written into the calculation tabs, recalculating
        would agree trivially and this whole test file would be checking that
        openpyxl can store a float.
        """
        from openpyxl import load_workbook
        from src.valuation.excel.workbook import SHEET_INPUTS

        book = load_workbook(exported, data_only=False)
        for sheet in book.worksheets:
            if sheet.title == SHEET_INPUTS:
                continue
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        pytest.fail(
                            f"{sheet.title}!{cell.coordinate} holds the literal "
                            f"{cell.value!r}; calculation cells must be formulas")

    def test_the_inputs_sheet_holds_no_formulas(self, exported):
        from openpyxl import load_workbook
        from src.valuation.excel.workbook import SHEET_INPUTS

        sheet = load_workbook(exported, data_only=False)[SHEET_INPUTS]
        formulas = [c.coordinate for row in sheet.iter_rows() for c in row
                    if isinstance(c.value, str) and c.value.startswith("=")]
        assert formulas == []


class TestReadback:
    def test_the_round_trip_is_exact(self, exported, spec_result):
        """Exact, not within a tolerance.

        A tolerance here would quietly absorb transcription errors as well as
        quantization. `to_spreadsheet` makes exactness achievable by writing
        what openpyxl will write anyway.
        """
        model, _ = spec_result
        quantized = read_inputs(exported, model.inputs)
        for field in LAYOUT.scalar_rows:
            assert getattr(quantized, field) == to_spreadsheet(
                getattr(model.inputs, field)), field

    def test_an_override_typed_into_the_sheet_reaches_python(
            self, exported, spec_result, tmp_path):
        """The half that makes the interface two-way."""
        from openpyxl import load_workbook
        from src.valuation.excel.workbook import SHEET_INPUTS

        model, _ = spec_result
        book = load_workbook(exported)
        book[SHEET_INPUTS].cell(LAYOUT.scalar_rows["terminal_growth"], 2,
                                Decimal("0.02"))
        edited = tmp_path / "edited.xlsx"
        book.save(edited)

        overridden = read_inputs(edited, model.inputs)
        assert overridden.terminal_growth == Decimal("0.02")
        result = discounted_cash_flow(overridden, Conventions.SPEC)
        assert result.share_price < discounted_cash_flow(
            read_inputs(exported, model.inputs), Conventions.SPEC).share_price

    def test_a_workbook_without_the_inputs_sheet_is_refused(self, tmp_path,
                                                            spec_result):
        from openpyxl import Workbook
        model, _ = spec_result
        path = tmp_path / "empty.xlsx"
        Workbook().save(path)
        with pytest.raises(ReadbackError, match="no 'Inputs' sheet"):
            read_inputs(path, model.inputs)

    def test_an_emptied_input_cell_is_refused_not_defaulted(self, exported,
                                                            spec_result,
                                                            tmp_path):
        from openpyxl import load_workbook
        from src.valuation.excel.workbook import SHEET_INPUTS

        model, _ = spec_result
        book = load_workbook(exported)
        book[SHEET_INPUTS].cell(LAYOUT.scalar_rows["beta"], 2).value = None
        path = tmp_path / "hole.xlsx"
        book.save(path)
        with pytest.raises(ReadbackError, match="beta"):
            read_inputs(path, model.inputs)


class TestTheExporterRefusesDefectiveModels:
    def test_a_defective_model_will_not_be_written(self, tmp_path, spec_result):
        """reader.py may READ a defective workbook; the writer may not emit one."""
        model, _ = spec_result
        as_built = discounted_cash_flow(model.inputs, model.conventions)
        with pytest.raises(ExportError, match="terminal_value_base"):
            write_workbook(as_built, tmp_path / "nope.xlsx")
