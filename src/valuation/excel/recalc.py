"""P2.5 — recalculate a workbook without Python's help (framework 4.12, C11).

C11 asserts that Excel's arithmetic and Python's agree. To assert it, something
other than Python has to do the arithmetic, and this module is the adapter for
whatever that something is.

Two engines
-----------
`formulas` is a pure-Python implementation of Excel's formula language. It
installs with pip, runs in CI, and needs no binary. It is also, unavoidably, a
Python program: reconciling our Python against somebody else's Python is
weaker evidence than reconciling against a real spreadsheet engine, and the
tolerance it satisfies should be read with that in mind.

LibreOffice is a real spreadsheet engine and therefore the stronger check, at
the cost of a large dependency that is awkward in CI. It is used when present.

Absence is an error, not a skip
-------------------------------
A skipped reconciliation is a gate that has quietly stopped running while
still reporting green, which is precisely the "warning by another name" that
framework 4.6 rejects. So a missing engine RAISES.

Opting out is possible and must be deliberate: `QC_RECALC_ENGINE=none`. The
report then states that the reconciliation did not run, because a workbook
that was never reconciled and one that reconciled cleanly are different
things and must not print the same.
"""
import os
import re
import shutil
import subprocess
import tempfile
from decimal import Decimal
from pathlib import Path

from src.valuation.money import from_spreadsheet

ENGINE_FORMULAS = "formulas"
ENGINE_LIBREOFFICE = "libreoffice"
ENGINE_NONE = "none"

ENGINE_ENV = "QC_RECALC_ENGINE"

_LIBREOFFICE_BINARIES = ("soffice", "libreoffice")

#: C11 tolerance, relative. Excel computes in binary64 (~15-17 significant
#: digits) and this engine in 50-digit Decimal, so agreement to 1e-9 is many
#: orders looser than the noise floor and many orders tighter than any
#: threshold in framework 4 -- a divergence this large is a formula that
#: disagrees, never a rounding artefact.
C11_TOLERANCE = Decimal("1e-9")


class RecalculationUnavailable(RuntimeError):
    """No engine can recalculate the workbook, and C11 cannot be skipped."""


def libreoffice_binary() -> str | None:
    for name in _LIBREOFFICE_BINARIES:
        found = shutil.which(name)
        if found:
            return found
    return None


def available_engines() -> tuple[str, ...]:
    engines = []
    try:
        import formulas  # noqa: F401
        engines.append(ENGINE_FORMULAS)
    except ImportError:
        pass
    if libreoffice_binary():
        engines.append(ENGINE_LIBREOFFICE)
    return tuple(engines)


def selected_engine() -> str:
    """The engine to use, honouring the environment, or a named failure."""
    requested = os.environ.get(ENGINE_ENV, "").strip().lower()
    if requested == ENGINE_NONE:
        return ENGINE_NONE

    available = available_engines()
    if requested:
        if requested not in available:
            raise RecalculationUnavailable(
                f"{ENGINE_ENV}={requested!r} but that engine is not available "
                f"(found: {available or 'none'})")
        return requested

    if not available:
        raise RecalculationUnavailable(
            "no recalculation engine is available, so the Excel/Python "
            "reconciliation (C11, framework 4.12) cannot run. Install one with "
            "`pip install formulas`, or install LibreOffice. Setting "
            f"{ENGINE_ENV}=none disables the check deliberately and makes the "
            "report say so -- a reconciliation that did not run and one that "
            "passed are different things.")
    # LibreOffice is the stronger evidence; prefer it when both are present.
    return (ENGINE_LIBREOFFICE if ENGINE_LIBREOFFICE in available
            else available[0])


def _normalise(key: str) -> str:
    """`'[book.xlsx]SHEET1'!B27` -> `SHEET1!B27`."""
    match = re.search(r"\]([^']+)'!(\$?[A-Z]+\$?\d+)$", key)
    if not match:
        return key.upper().replace("$", "")
    return f"{match.group(1)}!{match.group(2)}".upper().replace("$", "")


def _scalar(value):
    """`formulas` returns cells as nested arrays; take the single value."""
    while hasattr(value, "__len__") and not isinstance(value, str):
        if len(value) != 1:
            return None
        value = value[0]
    return value


def _recalc_formulas(path: Path) -> dict[str, Decimal]:
    import formulas as formulas_lib

    solution = formulas_lib.ExcelModel().loads(str(path)).finish().calculate()
    values: dict[str, Decimal] = {}
    for key, cell in solution.items():
        if "!" not in key:
            continue
        raw = _scalar(getattr(cell, "value", cell))
        if isinstance(raw, bool) or raw is None:
            continue
        try:
            values[_normalise(key)] = from_spreadsheet(raw)
        except Exception:                          # noqa: BLE001 - non-numeric cell
            continue
    return values


def _recalc_libreoffice(path: Path) -> dict[str, Decimal]:
    from openpyxl import load_workbook

    binary = libreoffice_binary()
    with tempfile.TemporaryDirectory() as outdir:
        subprocess.run(
            [binary, "--headless", "--norestore", "--convert-to", "xlsx",
             "--outdir", outdir, str(path)],
            check=True, capture_output=True, timeout=300)
        converted = Path(outdir) / path.with_suffix(".xlsx").name
        if not converted.exists():
            raise RecalculationUnavailable(
                f"LibreOffice produced no output for {path.name}")
        book = load_workbook(converted, data_only=True)
        values: dict[str, Decimal] = {}
        for sheet in book.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, (int, float)) and not isinstance(
                            cell.value, bool):
                        key = f"{sheet.title}!{cell.coordinate}".upper()
                        values[key] = from_spreadsheet(cell.value)
        return values


def recalculate(path: str | Path, engine: str | None = None
                ) -> dict[str, Decimal] | None:
    """Every numeric cell after recalculation, keyed `SHEET!REF` (upper case).

    Returns None when the check is deliberately disabled, so a caller can tell
    "did not run" from "ran and agreed" and report them differently.
    """
    engine = engine or selected_engine()
    if engine == ENGINE_NONE:
        return None
    path = Path(path)
    if engine == ENGINE_FORMULAS:
        return _recalc_formulas(path)
    if engine == ENGINE_LIBREOFFICE:
        return _recalc_libreoffice(path)
    raise RecalculationUnavailable(f"unknown engine {engine!r}")


def relative_difference(expected: Decimal, actual: Decimal) -> Decimal:
    """|actual - expected| / |expected|, and absolute difference at zero."""
    if expected == 0:
        return abs(actual)
    return abs(actual - expected) / abs(expected)


def agrees(expected: Decimal, actual: Decimal,
           tolerance: Decimal = C11_TOLERANCE) -> bool:
    return relative_difference(expected, actual) <= tolerance
