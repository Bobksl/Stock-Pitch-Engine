"""P2.5 — read overrides typed into `Inputs` back into Python (framework 4.12).

The half of the two-way interface that makes it two-way. Type a different
terminal growth rate into the workbook, read it back, and Python recomputes
from the value you typed rather than the one it exported.

Round-tripping is exact, and it is exact by construction rather than by luck:
`workbook.py` writes through `money.to_spreadsheet`, which quantizes to the
shortest decimal surviving a trip through binary64, so what comes back out is
what went in. Without that step a 50-digit stub fraction would return as
something slightly different and every round-trip assertion would need a
tolerance -- which would then quietly absorb real transcription errors too.

Only `Inputs` is read. The calculation sheets hold formulas, and their values
belong to whichever engine last recalculated the file; treating those as
authoritative would let a stale cached value overwrite a computed one.
"""
from dataclasses import replace
from pathlib import Path

from openpyxl import load_workbook

from src.valuation.excel.workbook import LAYOUT, SHEET_INPUTS, Layout
from src.valuation.inputs import ForecastYear, ValuationInputs
from src.valuation.money import from_spreadsheet


class ReadbackError(ValueError):
    """The workbook does not carry the inputs the layout expects."""


def read_inputs(path: str | Path, template: ValuationInputs,
                layout: Layout = LAYOUT) -> ValuationInputs:
    """The inputs as the workbook now holds them.

    `template` supplies what the sheet does not carry -- currency, unit, base
    period, provenance handles -- so an override changes the numbers and
    nothing else about the model's identity.
    """
    book = load_workbook(Path(path), data_only=True)
    if SHEET_INPUTS not in book.sheetnames:
        raise ReadbackError(
            f"no {SHEET_INPUTS!r} sheet; sheets are {book.sheetnames}")
    sheet = book[SHEET_INPUTS]

    def scalar(field: str):
        cell = sheet.cell(layout.scalar_rows[field], 2)
        if cell.value is None:
            raise ReadbackError(
                f"{SHEET_INPUTS}!B{layout.scalar_rows[field]} ({field}) is empty")
        return from_spreadsheet(cell.value)

    periods = []
    for index in range(len(template.forecast)):
        column = layout.first_period_column + index
        label = sheet.cell(layout.period_row, column).value
        if label is None:
            raise ReadbackError(
                f"{SHEET_INPUTS} row {layout.period_row} has no period in "
                f"column {column}: the workbook carries fewer forecast periods "
                f"than the model it is being read into")

        def driver(field: str, column=column):
            cell = sheet.cell(layout.driver_rows[field], column)
            if cell.value is None:
                raise ReadbackError(
                    f"{SHEET_INPUTS} row {layout.driver_rows[field]} ({field}) "
                    f"is empty for period {label}")
            return from_spreadsheet(cell.value)

        periods.append(ForecastYear(
            period=int(label),
            revenue_growth=driver("revenue_growth"),
            ebitda_margin=driver("ebitda_margin"),
            depreciation=driver("depreciation"),
            capex=driver("capex"),
            change_in_nwc=driver("change_in_nwc"),
            gross_margin=template.forecast[index].gross_margin,
        ))

    return replace(
        template,
        forecast=tuple(periods),
        **{field: scalar(field) for field in layout.scalar_rows})
