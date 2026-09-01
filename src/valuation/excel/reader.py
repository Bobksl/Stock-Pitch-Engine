"""P2.2 — read a foreign valuation workbook: its inputs, and its conventions.

A model in the wild cannot be parsed generically -- there is no schema, and
every analyst lays out a DCF differently -- so the layout is supplied as a
`CellMap`. What stays constant is the reader's job: read the literals the
modeller typed, and read the *formulas* to determine which conventions the
workbook actually implements.

Reading formulas matters more than it sounds. Three of the eight defects the
Phase 2 exit criterion requires (1, 6, 7) leave no trace in any value. A
workbook that builds terminal value from the already-discounted final cash
flow produces a perfectly ordinary-looking number; the only evidence that
anything is wrong is that the formula names the discounted row. Recomputing
from the inputs alone would reproduce the defect faithfully and report
nothing -- which is precisely how the error survived review in the first place.

This module answers "what does this workbook do". Mapping that to rule
findings with classes and spec references is `audit.py`, at P2.6.
"""
from dataclasses import dataclass
from decimal import Decimal
import re

from openpyxl import load_workbook

from src.valuation.inputs import (
    EQUITY_WEIGHT_MARKET_CAP,
    EQUITY_WEIGHT_MARKET_CAP_LESS_DEBT,
    STUB_FULL_YEAR_AT_STUB_FACTOR,
    STUB_PRORATE_CASH_FLOW,
    TV_FROM_DISCOUNTED_UFCF,
    TV_FROM_UNDISCOUNTED_UFCF,
    Conventions,
    ForecastYear,
    ValuationInputs,
)
from src.valuation.money import D, from_spreadsheet


#: Valuation methods and components a workbook may contain (framework 4.2,
#: 4.7, 4.10). A CellMap declares which are present, because for a foreign
#: workbook there is no reliable way to detect a comps tab or a scenario block
#: by inspection -- the same reason every other field of a CellMap is declared
#: rather than sniffed. Absence is the default, so an undescribed workbook is
#: reported as a single-method model rather than silently assumed complete.
COMPONENT_DCF = "dcf"
COMPONENT_COMPS = "comparable_companies"
COMPONENT_REVERSE_DCF = "reverse_dcf"
COMPONENT_SCENARIOS = "scenarios"
COMPONENT_SENSITIVITY = "sensitivity"

VALUATION_COMPONENTS = (COMPONENT_DCF, COMPONENT_COMPS, COMPONENT_REVERSE_DCF,
                        COMPONENT_SCENARIOS, COMPONENT_SENSITIVITY)


class WorkbookError(ValueError):
    """The workbook does not match the supplied cell map."""


@dataclass(frozen=True)
class CellMap:
    """Where the quantities live in one particular workbook's layout."""

    sheet: str
    currency: str
    unit: str

    # Scalars, as A1-style references
    risk_free_rate: str
    equity_risk_premium: str
    beta: str
    cost_of_debt: str
    tax_rate: str
    market_capitalisation: str      # the literal, wherever it is stated
    gross_debt: str
    terminal_growth: str
    total_debt: str
    cash_and_equivalents: str
    shares_outstanding: str

    # Rows. `columns` gives the forecast columns in period order; `base_revenue`
    # is the last actual, one column to their left.
    columns: tuple[str, ...]
    base_period: int
    base_revenue: str
    row_revenue_growth: int
    row_ebitda_margin: int
    row_depreciation: int
    row_capex: int
    row_change_in_nwc: int
    row_gross_margin: int | None

    # Formula cells inspected to determine conventions
    cell_equity_weight: str         # the formula stating equity value
    cell_terminal_value: str        # the formula stating terminal value
    row_ufcf: int
    row_discount_factor: int
    row_discounted_ufcf: int

    #: Scalars stated in trillions where the forecast is in billions. Scale is
    #: read, never inferred (framework 6.4), so a workbook that mixes them
    #: declares which cells are which rather than leaving it to a heuristic.
    trillion_scalars: tuple[str, ...] = ()

    #: A gross-profit row, where the workbook models one. Its only use here is
    #: to check whether anything references it (see WorkbookModel.dead_rows).
    row_gross_profit: int | None = None

    #: Which valuation components this workbook contains. DCF alone by default.
    components: frozenset[str] = frozenset({COMPONENT_DCF})


#: The layout of tests/fixtures/tsmc_model.xlsx. Data, not code: a second
#: workbook is a second CellMap, not a second reader.
TSMC_CELL_MAP = CellMap(
    sheet="Sheet1", currency="TWD", unit="billion",
    risk_free_rate="B2", equity_risk_premium="B3", beta="B1",
    cost_of_debt="B11", tax_rate="B12",
    market_capitalisation="B6", gross_debt="B5",
    terminal_growth="B17", total_debt="B23",
    cash_and_equivalents="B24", shares_outstanding="B26",
    columns=("I", "J", "K", "L", "M", "N"),
    base_period=2024, base_revenue="H3",
    row_revenue_growth=4, row_ebitda_margin=8, row_depreciation=9,
    row_capex=12, row_change_in_nwc=13, row_gross_margin=6,
    cell_equity_weight="B6", cell_terminal_value="B18",
    row_ufcf=14, row_discount_factor=15, row_discounted_ufcf=16,
    trillion_scalars=("market_capitalisation", "gross_debt"),
    row_gross_profit=5,
    # DCF only: no comps tab, no reverse DCF, no scenarios, no sensitivity
    # table. "Assign 100% weight on DCF approach", in the model's own words.
    components=frozenset({COMPONENT_DCF}),
)

_TRILLION_IN_BILLIONS = D(1000)


@dataclass(frozen=True)
class WorkbookModel:
    """A foreign model, read: what it asserted and how it combines it."""

    inputs: ValuationInputs
    conventions: Conventions
    formulas: dict[str, str]        # cell reference -> formula, for the audit
    published_price: Decimal | None
    cell_map: CellMap = TSMC_CELL_MAP

    def formula(self, ref: str) -> str:
        """The formula at a cell, for quoting as evidence in a finding."""
        return self.formulas.get(ref.upper(), "")

    @property
    def referenced_cells(self) -> set[str]:
        """Every cell named by any formula in the workbook."""
        return {ref for formula in self.formulas.values()
                for ref in _refs(formula)}

    def row_is_unreferenced(self, row: int) -> bool:
        """Whether no formula outside `row` reads any cell of it.

        A row computed for every period and read by nothing is a dead line: it
        costs nothing today and silently diverges from the live model the
        moment either is edited. The fixture's gross-profit row is one.
        """
        referenced = {ref for coordinate, formula in self.formulas.items()
                      if not coordinate[1:].isdigit() or int(coordinate[1:]) != row
                      for ref in _refs(formula)}
        return not any(f"{col}{row}" in referenced
                       for col in self.cell_map.columns)

    @property
    def dead_rows(self) -> tuple[int, ...]:
        """Declared rows that the workbook computes and never reads."""
        candidates = (self.cell_map.row_gross_profit,)
        return tuple(row for row in candidates
                     if row is not None and self.row_is_unreferenced(row))


def _literal(values, ref: str) -> Decimal:
    """A typed-in number, read from the cached value of the cell."""
    cell = values[ref]
    if cell.value is None:
        raise WorkbookError(f"{ref} is empty")
    return from_spreadsheet(cell.value)


def _formula(formulas, ref: str) -> str:
    raw = formulas[ref].value
    return raw if isinstance(raw, str) and raw.startswith("=") else ""


def _leading_literal(formulas, ref: str) -> Decimal:
    """The first numeric literal in a cell, e.g. 36.31 in '=36.31-B5'.

    A hardcode buried in a formula is still a hardcode, and reading it is how
    a workbook that never states market capitalisation in a cell of its own
    still yields one.
    """
    formula = _formula(formulas, ref)
    if not formula:
        return from_spreadsheet(formulas[ref].value)
    match = re.search(r"-?\d+(?:\.\d+)?", formula)
    if not match:
        raise WorkbookError(f"{ref} holds no numeric literal: {formula}")
    return D(match.group(0))


def _refs(formula: str) -> set[str]:
    """Cell references named by a formula, absolute markers stripped."""
    return {m.replace("$", "").upper()
            for m in re.findall(r"\$?[A-Z]{1,3}\$?\d{1,7}", formula.upper())}


def _stub_fraction(formula: str) -> Decimal:
    """The exponent of a first-period discount factor: 1/6 from POWER(1+r,1/6).

    Returns 1 when the formula carries no fraction, i.e. a full first period.
    """
    match = re.search(r",\s*(\d+)\s*/\s*(\d+)\s*\)", formula)
    if not match:
        return D(1)
    return D(match.group(1)) / D(match.group(2))


def read_model(path, cell_map: CellMap = TSMC_CELL_MAP,
               published_price_cell: str | None = None) -> WorkbookModel:
    """Read a workbook into declared inputs plus the conventions it implements."""
    formulas = load_workbook(path, data_only=False)[cell_map.sheet]
    values = load_workbook(path, data_only=True)[cell_map.sheet]
    m = cell_map

    def scaled(name: str, amount: Decimal) -> Decimal:
        return amount * _TRILLION_IN_BILLIONS if name in m.trillion_scalars else amount

    # The equity cell may state market capitalisation outright or bury it in a
    # formula netting debt out of it; either way the literal is the market cap.
    market_cap = _leading_literal(formulas, m.market_capitalisation)
    gross_debt = _literal(values, m.gross_debt)

    forecast = tuple(
        ForecastYear(
            period=m.base_period + 1 + offset,
            revenue_growth=_literal(values, f"{col}{m.row_revenue_growth}"),
            ebitda_margin=_literal(values, f"{col}{m.row_ebitda_margin}"),
            depreciation=_literal(values, f"{col}{m.row_depreciation}"),
            capex=_literal(values, f"{col}{m.row_capex}"),
            change_in_nwc=_literal(values, f"{col}{m.row_change_in_nwc}"),
            gross_margin=(_literal(values, f"{col}{m.row_gross_margin}")
                          if m.row_gross_margin else None),
        )
        for offset, col in enumerate(m.columns))

    first_col, last_col = m.columns[0], m.columns[-1]
    stub = _stub_fraction(_formula(formulas, f"{first_col}{m.row_discount_factor}"))

    inputs = ValuationInputs(
        currency=m.currency, unit=m.unit,
        base_period=m.base_period, base_revenue=_literal(values, m.base_revenue),
        forecast=forecast,
        risk_free_rate=_literal(values, m.risk_free_rate),
        equity_risk_premium=_literal(values, m.equity_risk_premium),
        beta=_literal(values, m.beta),
        cost_of_debt=_literal(values, m.cost_of_debt),
        tax_rate=_literal(values, m.tax_rate),
        market_capitalisation=scaled("market_capitalisation", market_cap),
        gross_debt=scaled("gross_debt", gross_debt),
        terminal_growth=_literal(values, m.terminal_growth),
        total_debt=_literal(values, m.total_debt),
        cash_and_equivalents=_literal(values, m.cash_and_equivalents),
        shares_outstanding=_literal(values, m.shares_outstanding),
        stub_fraction=stub,
        # A foreign workbook declares nothing. That absence is defects 4 and 5.
        provenance={},
    )

    return WorkbookModel(
        inputs=inputs,
        conventions=_conventions(formulas, m, first_col, last_col, stub),
        formulas={cell.coordinate: cell.value
                  for row in formulas.iter_rows() for cell in row
                  if isinstance(cell.value, str) and cell.value.startswith("=")},
        published_price=(_literal(values, published_price_cell)
                         if published_price_cell else None),
        cell_map=m,
    )


def _conventions(formulas, m: CellMap, first_col: str, last_col: str,
                 stub: Decimal) -> Conventions:
    """Which of the enumerated conventions this workbook's formulas implement."""
    tv_formula = _formula(formulas, m.cell_terminal_value)
    tv_refs = _refs(tv_formula)
    discounted_final = f"{last_col}{m.row_discounted_ufcf}"
    ufcf_final = f"{last_col}{m.row_ufcf}"
    if discounted_final in tv_refs:
        tv_base = TV_FROM_DISCOUNTED_UFCF
    elif ufcf_final in tv_refs:
        tv_base = TV_FROM_UNDISCOUNTED_UFCF
    else:
        raise WorkbookError(
            f"terminal value at {m.cell_terminal_value} references neither "
            f"{ufcf_final} nor {discounted_final}: {tv_formula}")

    equity_formula = _formula(formulas, m.cell_equity_weight)
    nets_debt_out = ("-" in equity_formula
                     and m.gross_debt.replace("$", "").upper() in _refs(equity_formula))

    # A stub factor applied to an untouched full-year cash flow: the discounted
    # cell multiplies the whole UFCF row rather than a pro-rated slice of it.
    first_discounted = _formula(formulas, f"{first_col}{m.row_discounted_ufcf}")
    full_year_at_stub = (stub != 1
                         and f"{first_col}{m.row_ufcf}" in _refs(first_discounted))

    return Conventions(
        terminal_value_base=tv_base,
        equity_weight_basis=(EQUITY_WEIGHT_MARKET_CAP_LESS_DEBT if nets_debt_out
                             else EQUITY_WEIGHT_MARKET_CAP),
        stub_policy=(STUB_FULL_YEAR_AT_STUB_FACTOR if full_year_at_stub
                     else STUB_PRORATE_CASH_FLOW),
    )
