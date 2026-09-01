"""P2.5 — export the valuation as a workbook of live formulas (framework 4.12).

Python owns calculation. The workbook is a two-way interface, not a report:
you type in `Inputs`, everything else recalculates, and `readback.py` reads
your overrides back into Python.

**No calculation cell holds a pasted value.** Every cell outside `Inputs` is a
formula referencing other cells. This is the property that makes C11 mean
anything at all -- if the exporter wrote Python's numbers into the calculation
tabs, recalculating the workbook would agree with Python trivially and the
reconciliation would be a test of openpyxl's ability to store a float. Writing
formulas instead means Excel genuinely recomputes the model from the inputs,
and agreement is evidence that two independent implementations concur.

`LAYOUT` is the single source of truth for where things live. The writer, the
readback and the reconciliation all address cells through it, so moving a row
is one edit rather than three that must be kept in step -- which is the same
class of bug C11 exists to catch, one level up.

Compliant models only
---------------------
The exporter refuses to write a model whose conventions differ from
`Conventions.SPEC`. It can *read* a defective workbook (that is `reader.py`'s
job and the whole point of the audit), but emitting formulas that implement a
terminal value built from an already-discounted cash flow would be shipping
the defect, and a file that leaves this repository should not contain one.
"""
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.valuation.dcf import DcfResult
from src.valuation.excel.formulas import (
    formula,
    max_range,
    min_range,
    sum_range,
)
from src.valuation.inputs import Conventions
from src.valuation.money import to_spreadsheet

SHEET_INPUTS = "Inputs"
SHEET_WACC = "WACC"
SHEET_MODEL = "Model"
SHEET_DCF = "DCF"
SHEET_COMPS = "Comps"
SHEET_TARGET = "Target"
SHEET_SCENARIOS = "Scenarios"
SHEET_SENSITIVITY = "Sensitivity"
SHEET_REVERSE = "Reverse DCF"
SHEET_SUMMARY = "Summary"

#: Every tab framework 4.12 names is now implemented.
PENDING_SHEETS: tuple[str, ...] = ()


class ExportError(ValueError):
    """The model cannot be written as a compliant workbook."""


@dataclass(frozen=True)
class Layout:
    """Where every quantity lives. Addressed by the writer, readback and C11."""

    #: Inputs sheet: field name -> row. Values sit in column B.
    scalar_rows: dict[str, int]
    #: Inputs sheet: driver name -> row, one column per forecast period.
    driver_rows: dict[str, int]
    #: Inputs sheet: reverse-DCF quantities. Solved in Python and written here
    #: as literals so the workbook can CHECK the solution rather than restate
    #: it -- Excel recomputes the price from the implied assumption, and it
    #: must come back equal to the market price.
    reverse_rows: dict[str, int] = field(default_factory=dict)
    #: Inputs sheet: the target-side quantities a comps valuation needs.
    comps_rows: dict[str, int] = field(default_factory=dict)
    #: Inputs sheet: per-peer rows, one column per peer.
    peer_rows: dict[str, int] = field(default_factory=dict)
    #: Inputs sheet: the price-target block (4.9).
    target_rows: dict[str, int] = field(default_factory=dict)
    #: Inputs sheet: the scenario block (4.10). One column per scenario.
    scenario_rows: dict[str, int] = field(default_factory=dict)
    #: Inputs sheet: the two sensitivity axes (4.10). One column per point.
    axis_rows: dict[str, int] = field(default_factory=dict)
    #: First forecast column on every sheet that has periods.
    first_period_column: int = 3            # C
    #: Inputs sheet row carrying the period labels.
    period_row: int = 26

    # WACC sheet rows
    wacc_cost_of_equity: int = 3
    wacc_after_tax_debt: int = 4
    wacc_equity_value: int = 5
    wacc_debt_value: int = 6
    wacc_debt_weight: int = 7
    wacc_equity_weight: int = 8
    wacc_value: int = 9

    # Model sheet rows
    model_period: int = 3
    model_revenue: int = 4
    model_ebitda: int = 5
    model_depreciation: int = 6
    model_ebit: int = 7
    model_tax: int = 8
    model_capex: int = 9
    model_nwc: int = 10
    model_ufcf: int = 11

    # DCF sheet rows
    dcf_period: int = 3
    dcf_exponent: int = 4
    dcf_factor: int = 5
    dcf_flow: int = 6
    dcf_present_value: int = 7
    dcf_pv_forecast: int = 9
    dcf_terminal_value: int = 10
    dcf_pv_terminal: int = 11
    dcf_enterprise_value: int = 12
    dcf_equity_value: int = 13
    dcf_share_price: int = 14

    # Reverse DCF sheet rows (4.7)
    reverse_market_price: int = 3
    reverse_implied_growth: int = 4
    reverse_pv_forecast: int = 6
    reverse_terminal_value: int = 7
    reverse_pv_terminal: int = 8
    reverse_enterprise_value: int = 9
    reverse_equity_value: int = 10
    reverse_share_price: int = 11
    reverse_residual: int = 12

    # Comps sheet rows (4.8). One column per peer.
    comps_peer: int = 3
    comps_multiple: int = 4
    comps_growth_adjusted: int = 5
    comps_implied_multiple: int = 6
    comps_implied_numerator: int = 7
    comps_implied_equity: int = 8
    comps_implied_price: int = 9
    comps_upside: int = 10
    comps_range_low: int = 12
    comps_range_high: int = 13
    comps_spread: int = 14

    # Target sheet rows (4.9)
    target_shares: int = 3
    target_base: int = 5
    target_after_rerating: int = 6
    target_after_revision: int = 7
    target_price: int = 8
    target_rerating: int = 10
    target_revision: int = 11
    target_roll_forward: int = 12
    target_total: int = 13
    target_upside: int = 14
    target_identity: int = 15

    # Scenarios sheet rows (4.10). One column per scenario.
    scenario_name: int = 3
    scenario_probability: int = 4
    scenario_price: int = 5
    scenario_contribution: int = 6
    scenario_upside: int = 7
    scenario_weight_check: int = 9
    scenario_expected_value: int = 10

    # Sensitivity sheet (4.10). The discount-factor block runs from
    # `sensitivity_factor_first` for one row per forecast period, so the rows
    # below it depend on the horizon and are computed rather than fixed.
    sensitivity_wacc: int = 3
    sensitivity_factor_first: int = 4

    def sensitivity_pv(self, periods: int) -> int:
        return self.sensitivity_factor_first + periods

    def sensitivity_header(self, periods: int) -> int:
        return self.sensitivity_pv(periods) + 2

    def sensitivity_grid_first(self, periods: int) -> int:
        return self.sensitivity_header(periods) + 1

    def column(self, index: int) -> str:
        return get_column_letter(self.first_period_column + index)


LAYOUT = Layout(
    scalar_rows={
        "risk_free_rate": 4,
        "equity_risk_premium": 5,
        "beta": 6,
        "cost_of_debt": 7,
        "tax_rate": 8,
        "market_capitalisation": 11,
        "gross_debt": 12,
        "terminal_growth": 15,
        "total_debt": 18,
        "cash_and_equivalents": 19,
        "shares_outstanding": 20,
        "stub_fraction": 21,
        "base_revenue": 24,
    },
    driver_rows={
        "revenue_growth": 27,
        "ebitda_margin": 28,
        "depreciation": 29,
        "capex": 30,
        "change_in_nwc": 31,
    },
    reverse_rows={
        "market_price": 34,
        "implied_terminal_growth": 35,
    },
    comps_rows={
        "target_metric_value": 39,
        "target_growth": 40,
        "net_debt": 41,
        "current_price": 42,
    },
    peer_rows={
        "ticker": 45,
        "numerator_value": 46,
        "metric_value": 47,
        "growth": 48,
    },
    scenario_rows={
        "name": 67,
        "probability": 68,
        "price": 69,
    },
    axis_rows={
        "terminal_growth": 72,
        "wacc": 73,
    },
    target_rows={
        "multiple": 52,
        "forward_metric": 53,
        "cash": 54,
        "debt": 55,
        "basic_shares": 56,
        "options": 57,
        "restricted_units": 58,
        "offering": 59,
        "current_price": 60,
        "prior_multiple": 62,
        "prior_metric": 63,
        "revised_metric": 64,
    },
)

_SCALAR_LABELS = {
    "risk_free_rate": "Risk-free rate",
    "equity_risk_premium": "Equity risk premium",
    "beta": "Beta",
    "cost_of_debt": "Cost of debt (marginal)",
    "tax_rate": "Tax rate (marginal)",
    "market_capitalisation": "Market capitalisation",
    "gross_debt": "Gross debt",
    "terminal_growth": "Terminal growth (nominal)",
    "total_debt": "Total debt (signed)",
    "cash_and_equivalents": "Cash and equivalents",
    "shares_outstanding": "Shares outstanding (fully diluted)",
    "stub_fraction": "Stub fraction of first period",
    "base_revenue": "Base period revenue",
}

_REVERSE_LABELS = {
    "market_price": "Market price today",
    "implied_terminal_growth": "Implied terminal growth (solved in Python)",
}

_COMPS_LABELS = {
    "target_metric_value": "Target forward metric (the denominator)",
    "target_growth": "Target forward growth",
    "net_debt": "Net debt (signed, debt less cash)",
    "current_price": "Current share price",
}

_PEER_LABELS = {
    "ticker": "Peer",
    "numerator_value": "Peer numerator (EV or equity value)",
    "metric_value": "Peer forward metric",
    "growth": "Peer forward growth",
}

_TARGET_LABELS = {
    "multiple": "Target multiple",
    "forward_metric": "Forward metric at the valuation date",
    "cash": "Cash (pro-forma)",
    "debt": "Debt (pro-forma)",
    "basic_shares": "Basic shares",
    "options": "Options",
    "restricted_units": "Restricted units",
    "offering": "Offering shares (priced, unsettled)",
    "current_price": "Current share price",
    "prior_multiple": "Multiple at the valuation date",
    "prior_metric": "Same-period metric at the valuation date",
    "revised_metric": "Same-period metric after revision",
}

_DRIVER_LABELS = {
    "revenue_growth": "Revenue growth",
    "ebitda_margin": "EBITDA margin",
    "depreciation": "Depreciation and amortisation",
    "capex": "Capital expenditure (signed)",
    "change_in_nwc": "Change in net working capital (signed)",
}

_INPUT_FILL = PatternFill("solid", fgColor="DCE6F1")   # typed here, and only here
_HEADING = Font(bold=True)
_TITLE = Font(bold=True, size=12)


def _title(sheet, text: str, subtitle: str = "") -> None:
    sheet["A1"] = text
    sheet["A1"].font = _TITLE
    if subtitle:
        sheet["A2"] = subtitle


def _label(sheet, row: int, text: str, bold: bool = False) -> None:
    cell = sheet.cell(row, 1, text)
    if bold:
        cell.font = _HEADING


def _write_inputs(sheet, result: DcfResult, layout: Layout) -> None:
    inputs = result.inputs
    _title(sheet, f"Inputs — {inputs.currency} {inputs.unit}",
           "The only sheet you type in. Every other cell is a live formula.")

    for row, text in ((3, "Cost of capital (4.3)"),
                      (10, "Capital structure — equity weight is market cap (4.3)"),
                      (14, "Terminal value (4.6)"),
                      (17, "Equity bridge (4.9)"),
                      (23, "Explicit forecast — the Section 3 bridge (4.5)")):
        _label(sheet, row, text, bold=True)

    for name, row in layout.scalar_rows.items():
        _label(sheet, row, _SCALAR_LABELS[name])
        cell = sheet.cell(row, 2, to_spreadsheet(getattr(inputs, name)))
        cell.fill = _INPUT_FILL

    _label(sheet, layout.period_row, "Period", bold=True)
    for index, year in enumerate(inputs.forecast):
        cell = sheet.cell(layout.period_row, layout.first_period_column + index,
                          year.period)
        cell.font = _HEADING
        cell.alignment = Alignment(horizontal="right")

    for name, row in layout.driver_rows.items():
        _label(sheet, row, _DRIVER_LABELS[name])
        for index, year in enumerate(inputs.forecast):
            cell = sheet.cell(row, layout.first_period_column + index,
                              to_spreadsheet(getattr(year, name)))
            cell.fill = _INPUT_FILL

    sheet.column_dimensions["A"].width = 42


def _write_reverse_inputs(sheet, reverse, layout) -> None:
    """The solved reverse-DCF quantities, written as literals to be checked."""
    _label(sheet, 33, "Reverse DCF (4.7) — solved in Python, checked in Excel",
           bold=True)
    values = {"market_price": reverse.market_price,
              "implied_terminal_growth": reverse.terminal_growth.value}
    for name, row in layout.reverse_rows.items():
        _label(sheet, row, _REVERSE_LABELS[name])
        cell = sheet.cell(row, 2, to_spreadsheet(values[name]))
        cell.fill = _INPUT_FILL


def _rev(layout: Layout, name: str) -> str:
    return f"{SHEET_INPUTS}!B{layout.reverse_rows[name]}"


def _cmp(layout: Layout, name: str, absolute: bool = True) -> str:
    row = layout.comps_rows[name]
    return f"{SHEET_INPUTS}!{'$B$' if absolute else 'B'}{row}"


def _peer(layout: Layout, name: str, index: int) -> str:
    return f"{SHEET_INPUTS}!{layout.column(index)}{layout.peer_rows[name]}"


def _inp(layout: Layout, name: str, absolute: bool = False) -> str:
    row = layout.scalar_rows[name]
    return f"{SHEET_INPUTS}!{'$B$' if absolute else 'B'}{row}"


def _driver(layout: Layout, name: str, index: int) -> str:
    return (f"{SHEET_INPUTS}!{layout.column(index)}"
            f"{layout.driver_rows[name]}")


def _write_wacc(sheet, layout: Layout) -> None:
    _title(sheet, "WACC — built from inputs, never copied (4.3)",
           "Equity weight is market capitalisation; debt is not netted out of it.")
    rows = {
        layout.wacc_cost_of_equity: (
            "Cost of equity (CAPM)",
            "=" + _inp(layout, "risk_free_rate")
            + f"+{_inp(layout, 'beta')}*{_inp(layout, 'equity_risk_premium')}"),
        layout.wacc_after_tax_debt: (
            "Cost of debt, after tax",
            formula("product", _inp(layout, "cost_of_debt"),
                    f"(1-{_inp(layout, 'tax_rate')})")),
        layout.wacc_equity_value: (
            "Equity value (market capitalisation)",
            f"={_inp(layout, 'market_capitalisation')}"),
        layout.wacc_debt_value: ("Debt value", f"={_inp(layout, 'gross_debt')}"),
        layout.wacc_debt_weight: (
            "D / (D + E)",
            formula("ratio", f"B{layout.wacc_debt_value}",
                    f"(B{layout.wacc_equity_value}+B{layout.wacc_debt_value})")),
        layout.wacc_equity_weight: (
            "E / (D + E)", formula("complement", f"B{layout.wacc_debt_weight}")),
        layout.wacc_value: (
            "WACC",
            formula("weighted_pair", f"B{layout.wacc_cost_of_equity}",
                    f"B{layout.wacc_equity_weight}",
                    f"B{layout.wacc_after_tax_debt}",
                    f"B{layout.wacc_debt_weight}")),
    }
    for row, (label, cell_formula) in rows.items():
        _label(sheet, row, label, bold=row == layout.wacc_value)
        sheet.cell(row, 2, cell_formula)
    sheet.column_dimensions["A"].width = 38


def _write_model(sheet, result: DcfResult, layout: Layout) -> None:
    _title(sheet, "Model — the explicit forecast (4.5)",
           "Drivers are the Section 3 bridge, not independent assumptions.")
    labels = {
        layout.model_period: "Period",
        layout.model_revenue: "Revenue",
        layout.model_ebitda: "EBITDA",
        layout.model_depreciation: "Depreciation and amortisation",
        layout.model_ebit: "EBIT",
        layout.model_tax: "Tax on EBIT",
        layout.model_capex: "Capital expenditure",
        layout.model_nwc: "Change in net working capital",
        layout.model_ufcf: "Unlevered free cash flow",
    }
    for row, text in labels.items():
        _label(sheet, row, text,
               bold=row in (layout.model_period, layout.model_ufcf))

    for index in range(len(result.inputs.forecast)):
        col = layout.column(index)
        set_ = lambda row, value: sheet.cell(  # noqa: E731 - local, one line
            row, layout.first_period_column + index, value)

        set_(layout.model_period,
             f"={SHEET_INPUTS}!{col}{layout.period_row}")
        previous = (_inp(layout, "base_revenue") if index == 0
                    else f"{layout.column(index - 1)}{layout.model_revenue}")
        set_(layout.model_revenue,
             formula("grow_by", previous, _driver(layout, "revenue_growth", index)))
        set_(layout.model_ebitda,
             formula("product", f"{col}{layout.model_revenue}",
                     _driver(layout, "ebitda_margin", index)))
        set_(layout.model_depreciation, f"={_driver(layout, 'depreciation', index)}")
        set_(layout.model_ebit,
             formula("difference", f"{col}{layout.model_ebitda}",
                     f"{col}{layout.model_depreciation}"))
        set_(layout.model_tax,
             formula("negated_product", f"{col}{layout.model_ebit}",
                     _inp(layout, "tax_rate", absolute=True)))
        set_(layout.model_capex, f"={_driver(layout, 'capex', index)}")
        set_(layout.model_nwc, f"={_driver(layout, 'change_in_nwc', index)}")
        set_(layout.model_ufcf,
             formula("sum", f"{col}{layout.model_ebitda}", f"{col}{layout.model_tax}",
                     f"{col}{layout.model_capex}", f"{col}{layout.model_nwc}"))

    sheet.column_dimensions["A"].width = 38


def _write_dcf(sheet, result: DcfResult, layout: Layout) -> None:
    _title(sheet, "DCF — discounting and the equity bridge (4.6, 4.9)",
           "Terminal value is built from the final UNDISCOUNTED cash flow.")
    periods = len(result.inputs.forecast)
    last = layout.column(periods - 1)
    wacc = f"{SHEET_WACC}!$B${layout.wacc_value}"

    for row, text in {layout.dcf_period: "Period",
                      layout.dcf_exponent: "Discount exponent (years)",
                      layout.dcf_factor: "Discount factor",
                      layout.dcf_flow: "Cash flow discounted",
                      layout.dcf_present_value: "Present value"}.items():
        _label(sheet, row, text, bold=row == layout.dcf_period)

    for index in range(periods):
        col = layout.column(index)
        set_ = lambda row, value: sheet.cell(  # noqa: E731
            row, layout.first_period_column + index, value)
        set_(layout.dcf_period, f"={SHEET_MODEL}!{col}{layout.model_period}")
        stub = _inp(layout, "stub_fraction", absolute=True)
        set_(layout.dcf_exponent,
             f"={stub}" if index == 0 else f"={stub}+{index}")
        set_(layout.dcf_factor,
             formula("discount_factor", wacc, f"{col}{layout.dcf_exponent}"))
        # The stub period contributes a pro-rated slice of its cash flow (4.5).
        set_(layout.dcf_flow,
             formula("product", f"{SHEET_MODEL}!{col}{layout.model_ufcf}", stub)
             if index == 0 else f"={SHEET_MODEL}!{col}{layout.model_ufcf}")
        set_(layout.dcf_present_value,
             formula("product", f"{col}{layout.dcf_flow}",
                     f"{col}{layout.dcf_factor}"))

    first = layout.column(0)
    totals = {
        layout.dcf_pv_forecast: (
            "PV of explicit forecast",
            sum_range(f"{first}{layout.dcf_present_value}",
                      f"{last}{layout.dcf_present_value}")),
        layout.dcf_terminal_value: (
            "Terminal value",
            formula("perpetuity", f"{SHEET_MODEL}!{last}{layout.model_ufcf}",
                    _inp(layout, "terminal_growth"), wacc)),
        layout.dcf_pv_terminal: (
            "PV of terminal value",
            formula("product", f"B{layout.dcf_terminal_value}",
                    f"{last}{layout.dcf_factor}")),
        layout.dcf_enterprise_value: (
            "Enterprise value",
            formula("sum", f"B{layout.dcf_pv_forecast}",
                    f"B{layout.dcf_pv_terminal}")),
        layout.dcf_equity_value: (
            "Equity value",
            formula("sum", f"B{layout.dcf_enterprise_value}",
                    _inp(layout, "total_debt"),
                    _inp(layout, "cash_and_equivalents"))),
        layout.dcf_share_price: (
            "Implied share price",
            formula("ratio", f"B{layout.dcf_equity_value}",
                    _inp(layout, "shares_outstanding"))),
    }
    for row, (label, cell_formula) in totals.items():
        _label(sheet, row, label, bold=row == layout.dcf_share_price)
        sheet.cell(row, 2, cell_formula)

    sheet.column_dimensions["A"].width = 38


def _write_scenario_inputs(sheet, scenarios, axes, layout: Layout) -> None:
    """The scenario block and the two sensitivity axes, as typed values."""
    if scenarios is not None:
        _label(sheet, 66, "Scenarios (4.10) — coarse weights, pillar-traced",
               bold=True)
        for name, row in layout.scenario_rows.items():
            _label(sheet, row, name.replace("_", " ").capitalize())
        for index, outcome in enumerate(scenarios.outcomes):
            column = layout.first_period_column + index
            cell = sheet.cell(layout.scenario_rows["name"], column,
                              outcome.scenario.name)
            cell.font = _HEADING
            for key, value in (("probability", outcome.scenario.probability),
                               ("price", outcome.result.share_price)):
                written = sheet.cell(layout.scenario_rows[key], column,
                                     to_spreadsheet(value))
                written.fill = _INPUT_FILL

    if axes is not None:
        _label(sheet, 71, "Sensitivity axes (4.10)", bold=True)
        for name, values in axes.items():
            row = layout.axis_rows[name]
            _label(sheet, row, name.replace("_", " ").capitalize())
            for index, value in enumerate(values):
                cell = sheet.cell(row, layout.first_period_column + index,
                                  to_spreadsheet(value))
                cell.fill = _INPUT_FILL


def _write_scenarios(sheet, layout: Layout, scenarios) -> None:
    """Weight the cases and check the weights (framework 4.10).

    Scenario prices arrive as declared inputs, the same arrangement the
    Reverse DCF tab uses: Python values each case by rebuilding the whole
    model against a different set of drivers, which is not something a single
    column of formulas can express. What Excel checks here is the part it can
    own outright -- that the weights sum to one and that the expected value is
    the weighted sum it claims to be.

    The weight check is on the face of the sheet rather than in a test. A
    probability-weighted expected value over weights that do not sum to one is
    not an expected value, and that failure should be visible to whoever opens
    the file.
    """
    _title(sheet, "Scenarios — probability-weighted (4.10)",
           "Weights are coarse by design. Nobody knows a thesis is 63% likely.")

    for row, text in ((layout.scenario_name, "Scenario"),
                      (layout.scenario_probability, "Probability"),
                      (layout.scenario_price, "Target price"),
                      (layout.scenario_contribution, "Contribution"),
                      (layout.scenario_upside, "Upside")):
        _label(sheet, row, text, bold=row == layout.scenario_name)

    count = len(scenarios.outcomes)
    for index in range(count):
        col = layout.column(index)
        source = f"{SHEET_INPUTS}!{col}"
        sheet.cell(layout.scenario_name, layout.first_period_column + index,
                   f"={source}{layout.scenario_rows['name']}")
        sheet.cell(layout.scenario_probability,
                   layout.first_period_column + index,
                   f"={source}{layout.scenario_rows['probability']}")
        sheet.cell(layout.scenario_price, layout.first_period_column + index,
                   f"={source}{layout.scenario_rows['price']}")
        sheet.cell(layout.scenario_contribution,
                   layout.first_period_column + index,
                   formula("product", f"{col}{layout.scenario_probability}",
                           f"{col}{layout.scenario_price}"))
        sheet.cell(layout.scenario_upside, layout.first_period_column + index,
                   formula("upside", f"{col}{layout.scenario_price}",
                           f"{SHEET_DCF}!$B${layout.dcf_share_price}"))

    first, last = layout.column(0), layout.column(count - 1)
    _label(sheet, layout.scenario_weight_check,
           "Weights (must be exactly 1)", bold=True)
    sheet.cell(layout.scenario_weight_check, 2,
               sum_range(f"{first}{layout.scenario_probability}",
                         f"{last}{layout.scenario_probability}"))
    _label(sheet, layout.scenario_expected_value,
           "Probability-weighted expected value", bold=True)
    sheet.cell(layout.scenario_expected_value, 2,
               sum_range(f"{first}{layout.scenario_contribution}",
                         f"{last}{layout.scenario_contribution}"))
    sheet.column_dimensions["A"].width = 36


def _write_sensitivity(sheet, layout: Layout, periods: int,
                       growth_points: int, wacc_points: int) -> None:
    """A genuine two-way grid: terminal growth against WACC (framework 4.10).

    Every cell is computed in Excel from first principles rather than pasted,
    which is why the sheet is built in two parts. Changing WACC changes every
    discount factor, so each WACC column first gets its own factor block and
    its own present value of the explicit forecast; the grid below then values
    the perpetuity at each growth rate against each column's WACC.

    The cash flows come from the DCF sheet's discounted-flow row, which is
    already stub-adjusted, so the stub convention is applied once in the model
    rather than repeated -- and repeated differently -- in twenty-five places.
    """
    _title(sheet, "Sensitivity — terminal growth against WACC (4.10)",
           "Each column rebuilds its own discount factors; nothing is pasted.")

    axis_wacc = f"{SHEET_INPUTS}!{{col}}{layout.axis_rows['wacc']}"
    axis_growth = f"{SHEET_INPUTS}!$C${layout.axis_rows['terminal_growth']}"
    stub = _inp(layout, "stub_fraction", absolute=True)
    pv_row = layout.sensitivity_pv(periods)

    _label(sheet, layout.sensitivity_wacc, "WACC", bold=True)
    for offset in range(periods):
        _label(sheet, layout.sensitivity_factor_first + offset,
               f"Discount factor, period {offset + 1}")
    _label(sheet, pv_row, "PV of explicit forecast", bold=True)

    for index in range(wacc_points):
        col = layout.column(index)
        wacc_cell = f"{col}{layout.sensitivity_wacc}"
        sheet.cell(layout.sensitivity_wacc, layout.first_period_column + index,
                   "=" + axis_wacc.format(col=col))
        for offset in range(periods):
            sheet.cell(layout.sensitivity_factor_first + offset,
                       layout.first_period_column + index,
                       formula("discount_factor",
                               f"{col}${layout.sensitivity_wacc}",
                               f"{stub}+{offset}" if offset else stub))
        terms = "+".join(
            f"{SHEET_DCF}!{layout.column(offset)}${layout.dcf_flow}"
            f"*{col}{layout.sensitivity_factor_first + offset}"
            for offset in range(periods))
        sheet.cell(pv_row, layout.first_period_column + index, f"={terms}")

    header_row = layout.sensitivity_header(periods)
    _label(sheet, header_row, "Terminal growth \\ WACC", bold=True)
    for index in range(wacc_points):
        col = layout.column(index)
        sheet.cell(header_row, layout.first_period_column + index,
                   f"={col}{layout.sensitivity_wacc}")

    last_flow = layout.column(periods - 1)
    for row_index in range(growth_points):
        row = layout.sensitivity_grid_first(periods) + row_index
        growth_cell = (f"{SHEET_INPUTS}!{layout.column(row_index)}"
                       f"${layout.axis_rows['terminal_growth']}")
        sheet.cell(row, 2, f"={growth_cell}")
        for column_index in range(wacc_points):
            col = layout.column(column_index)
            wacc_cell = f"{col}${layout.sensitivity_wacc}"
            terminal = (f"{SHEET_MODEL}!${last_flow}${layout.model_ufcf}"
                        f"*(1+$B{row})/({wacc_cell}-$B{row})")
            sheet.cell(row, layout.first_period_column + column_index,
                       f"=({col}${pv_row}+({terminal})"
                       f"*{col}{layout.sensitivity_factor_first + periods - 1}"
                       f"+{_inp(layout, 'total_debt', absolute=True)}"
                       f"+{_inp(layout, 'cash_and_equivalents', absolute=True)})"
                       f"/{_inp(layout, 'shares_outstanding', absolute=True)}")

    sheet.column_dimensions["A"].width = 30


def _write_target_inputs(sheet, target, layout: Layout) -> None:
    """The price-target block, as typed values (framework 4.9)."""
    _label(sheet, 51, "Price target (4.9) — fully diluted, source declared",
           bold=True)
    shares = target.shares
    values = {
        "multiple": target.multiple,
        "forward_metric": target.forward_metric,
        "cash": target.cash,
        "debt": target.debt,
        "basic_shares": shares.basic,
        "options": shares.options,
        "restricted_units": shares.restricted_units,
        "offering": shares.offering,
        "current_price": target.current_price,
        "prior_multiple": target.prior_multiple,
        "prior_metric": target.prior_metric,
        "revised_metric": (target.revised_metric
                           if target.revised_metric is not None
                           else target.prior_metric),
    }
    for name, row in layout.target_rows.items():
        _label(sheet, row, _TARGET_LABELS[name])
        value = values[name]
        if value is None:
            continue
        cell = sheet.cell(row, 2, to_spreadsheet(value))
        cell.fill = _INPUT_FILL


def _write_target(sheet, layout: Layout, target) -> None:
    """The bridge, and the decomposition that makes the target auditable.

    The last row is an identity check: the three components minus the actual
    price change, which must be zero. Excel computes it independently, so a
    decomposition that does not add up says so on the face of the workbook
    rather than in a test somebody has to run.
    """
    _title(sheet, "Target — price and return decomposition (4.9)",
           "Of the total upside, how much is re-rating, revision, roll-forward.")

    def tgt(name: str) -> str:
        return f"{SHEET_INPUTS}!$B${layout.target_rows[name]}"

    def price_at(multiple: str, metric: str) -> str:
        return formula("implied_price", multiple, metric, tgt("cash"),
                       tgt("debt"), f"$B${layout.target_shares}")

    rows = {
        layout.target_shares: (
            "Fully diluted shares",
            formula("sum", tgt("basic_shares"), tgt("options"),
                    tgt("restricted_units"), tgt("offering"))),
        layout.target_base: (
            "Price at the valuation date",
            price_at(tgt("prior_multiple"), tgt("prior_metric"))),
        layout.target_after_rerating: (
            "...after re-rating",
            price_at(tgt("multiple"), tgt("prior_metric"))),
        layout.target_after_revision: (
            "...after estimate revision",
            price_at(tgt("multiple"), tgt("revised_metric"))),
        layout.target_price: (
            "...after rolling forward = target",
            price_at(tgt("multiple"), tgt("forward_metric"))),
        layout.target_rerating: (
            "Re-rating",
            formula("difference", f"B{layout.target_after_rerating}",
                    f"B{layout.target_base}")),
        layout.target_revision: (
            "Estimate revision",
            formula("difference", f"B{layout.target_after_revision}",
                    f"B{layout.target_after_rerating}")),
        layout.target_roll_forward: (
            "Time roll-forward",
            formula("difference", f"B{layout.target_price}",
                    f"B{layout.target_after_revision}")),
        layout.target_total: (
            "Total price change",
            formula("sum", f"B{layout.target_rerating}",
                    f"B{layout.target_revision}",
                    f"B{layout.target_roll_forward}")),
        layout.target_upside: (
            "Upside",
            formula("upside", f"B{layout.target_price}", tgt("current_price"))),
        layout.target_identity: (
            "Identity check (must be zero)",
            # Parenthesised: difference() emits 'a-b', so an unbracketed
            # 'B8-B5' would render as B13-B8-B5 and quietly check nothing.
            formula("difference", f"B{layout.target_total}",
                    f"(B{layout.target_price}-B{layout.target_base})")),
    }
    for row, (text, cell_formula) in rows.items():
        _label(sheet, row, text,
               bold=row in (layout.target_price, layout.target_upside,
                            layout.target_identity))
        sheet.cell(row, 2, cell_formula)
    sheet.column_dimensions["A"].width = 40


def _write_comps_inputs(sheet, comps, layout: Layout) -> None:
    """The target-side quantities and the peer table, as typed values."""
    _label(sheet, 38, "Comparable companies (4.8) — same period, same source",
           bold=True)
    values = {"target_metric_value": comps.target_metric_value,
              "target_growth": comps.target_growth,
              "net_debt": comps.net_debt,
              "current_price": comps.current_price}
    for name, row in layout.comps_rows.items():
        _label(sheet, row, _COMPS_LABELS[name])
        value = values[name]
        if value is None:
            continue
        cell = sheet.cell(row, 2, to_spreadsheet(value))
        cell.fill = _INPUT_FILL

    _label(sheet, layout.peer_rows["ticker"], _PEER_LABELS["ticker"], bold=True)
    for name in ("numerator_value", "metric_value", "growth"):
        _label(sheet, layout.peer_rows[name], _PEER_LABELS[name])

    for index, peer in enumerate(comps.peers):
        column = layout.first_period_column + index
        ticker = sheet.cell(layout.peer_rows["ticker"], column, peer.ticker)
        ticker.font = _HEADING
        for name in ("numerator_value", "metric_value", "growth"):
            cell = sheet.cell(layout.peer_rows[name], column,
                              to_spreadsheet(getattr(peer, name)))
            cell.fill = _INPUT_FILL


def _write_comps(sheet, layout: Layout, comps) -> None:
    """Every peer anchor, computed live. The range is the deliverable.

    Framework 4.8 makes anchor disclosure mandatory because the choice of peer
    moves the headline more than the method does -- Reddit at IPO is 22-30%
    upside anchored to PINS and +8% anchored to SNAP, same inputs, same day.
    So this tab computes ALL of them side by side and puts the range and the
    spread on the bottom three rows, where a reader cannot take one column
    home without seeing the others.

    The regression (the primary normalisation method) is deliberately NOT
    here. It would need LINEST, whose support in the reconciliation engine is
    unverified, and an unreconciled formula on a calculation tab is precisely
    what C11 exists to prevent. It stays in Python and is reported there.
    """
    from src.valuation.comps import NUMERATOR_ENTERPRISE, label as multiple_label

    _title(sheet,
           f"Comps — {multiple_label(comps.numerator, comps.metric)}, "
           f"every anchor (4.8)",
           "Growth-adjusted multiples are a CROSS-CHECK. The range is the point.")

    for row, text in ((layout.comps_peer, "Peer"),
                      (layout.comps_multiple, "Peer multiple"),
                      (layout.comps_growth_adjusted, "Per point of growth"),
                      (layout.comps_implied_multiple, "Implied multiple"),
                      (layout.comps_implied_numerator, "Implied numerator"),
                      (layout.comps_implied_equity, "Implied equity value"),
                      (layout.comps_implied_price, "Implied price"),
                      (layout.comps_upside, "Upside")):
        _label(sheet, row, text,
               bold=row in (layout.comps_peer, layout.comps_implied_multiple))

    count = len(comps.peers)
    for index in range(count):
        col = layout.column(index)
        set_ = lambda row, value: sheet.cell(  # noqa: E731
            row, layout.first_period_column + index, value)

        set_(layout.comps_peer, f"={_peer(layout, 'ticker', index)}")
        set_(layout.comps_multiple,
             formula("ratio", _peer(layout, "numerator_value", index),
                     _peer(layout, "metric_value", index)))
        set_(layout.comps_growth_adjusted,
             formula("per_point", f"{col}{layout.comps_multiple}",
                     _peer(layout, "growth", index)))
        set_(layout.comps_implied_multiple,
             formula("times_points", f"{col}{layout.comps_growth_adjusted}",
                     _cmp(layout, "target_growth")))
        set_(layout.comps_implied_numerator,
             formula("product", f"{col}{layout.comps_implied_multiple}",
                     _cmp(layout, "target_metric_value")))
        set_(layout.comps_implied_equity,
             formula("difference", f"{col}{layout.comps_implied_numerator}",
                     _cmp(layout, "net_debt"))
             if comps.numerator == NUMERATOR_ENTERPRISE
             else f"={col}{layout.comps_implied_numerator}")
        set_(layout.comps_implied_price,
             formula("ratio", f"{col}{layout.comps_implied_equity}",
                     _inp(layout, "shares_outstanding", absolute=True)))
        set_(layout.comps_upside,
             formula("upside", f"{col}{layout.comps_implied_price}",
                     _cmp(layout, "current_price")))

    first, last = layout.column(0), layout.column(count - 1)
    row_of = layout.comps_implied_multiple
    totals = {
        layout.comps_range_low: (
            "Range low (implied multiple)",
            min_range(f"{first}{row_of}", f"{last}{row_of}")),
        layout.comps_range_high: (
            "Range high (implied multiple)",
            max_range(f"{first}{row_of}", f"{last}{row_of}")),
        layout.comps_spread: (
            "Factor on the choice of peer",
            formula("ratio", f"B{layout.comps_range_high}",
                    f"B{layout.comps_range_low}")),
    }
    for row, (text, cell_formula) in totals.items():
        _label(sheet, row, text, bold=row == layout.comps_spread)
        sheet.cell(row, 2, cell_formula)
    sheet.column_dimensions["A"].width = 38


def _write_reverse(sheet, layout: Layout, periods: int) -> None:
    """Recompute the price at the implied assumption; it must equal the market.

    The tab does not restate Python's answer -- it CHECKS it. Python solves for
    the terminal growth rate today's price implies; this sheet takes that rate
    as an input, rebuilds the valuation around it, and the residual on the last
    row is Excel's verdict on whether the solve was right. A pasted "implied
    growth: 2.35%" would assert the result; this interrogates it.

    Only the terminal branch is rebuilt. The explicit forecast's present value
    does not depend on g at all, so it is referenced from the DCF sheet rather
    than duplicated -- a second copy would be a second thing to drift.
    """
    _title(sheet, "Reverse DCF — what today's price already assumes (4.7)",
           "The residual on the last row is Excel checking Python's solve.")
    last = layout.column(periods - 1)
    wacc = f"{SHEET_WACC}!$B${layout.wacc_value}"
    growth = _rev(layout, "implied_terminal_growth")

    rows = {
        layout.reverse_market_price: (
            "Market price today", f"={_rev(layout, 'market_price')}"),
        layout.reverse_implied_growth: (
            "Implied terminal growth", f"={growth}"),
        layout.reverse_pv_forecast: (
            "PV of explicit forecast (independent of g)",
            f"={SHEET_DCF}!B{layout.dcf_pv_forecast}"),
        layout.reverse_terminal_value: (
            "Terminal value at the implied growth",
            formula("perpetuity", f"{SHEET_MODEL}!{last}{layout.model_ufcf}",
                    growth, wacc)),
        layout.reverse_pv_terminal: (
            "PV of terminal value",
            formula("product", f"B{layout.reverse_terminal_value}",
                    f"{SHEET_DCF}!{last}{layout.dcf_factor}")),
        layout.reverse_enterprise_value: (
            "Enterprise value",
            formula("sum", f"B{layout.reverse_pv_forecast}",
                    f"B{layout.reverse_pv_terminal}")),
        layout.reverse_equity_value: (
            "Equity value",
            formula("sum", f"B{layout.reverse_enterprise_value}",
                    _inp(layout, "total_debt"),
                    _inp(layout, "cash_and_equivalents"))),
        layout.reverse_share_price: (
            "Implied share price",
            formula("ratio", f"B{layout.reverse_equity_value}",
                    _inp(layout, "shares_outstanding"))),
        layout.reverse_residual: (
            "Residual against the market price",
            formula("difference", f"B{layout.reverse_share_price}",
                    f"B{layout.reverse_market_price}")),
    }
    for row, (label, cell_formula) in rows.items():
        _label(sheet, row, label,
               bold=row in (layout.reverse_share_price, layout.reverse_residual))
        sheet.cell(row, 2, cell_formula)
    sheet.column_dimensions["A"].width = 42


def _write_summary(sheet, result: DcfResult, layout: Layout) -> None:
    inputs = result.inputs
    _title(sheet, f"Summary — {inputs.currency} {inputs.unit}",
           "Every figure below is a reference. Nothing is typed on this sheet.")
    rows = {
        4: ("Implied share price", f"={SHEET_DCF}!B{layout.dcf_share_price}"),
        5: ("Enterprise value", f"={SHEET_DCF}!B{layout.dcf_enterprise_value}"),
        6: ("PV of explicit forecast", f"={SHEET_DCF}!B{layout.dcf_pv_forecast}"),
        7: ("PV of terminal value", f"={SHEET_DCF}!B{layout.dcf_pv_terminal}"),
        8: ("Terminal value share of EV",
            formula("ratio", f"{SHEET_DCF}!B{layout.dcf_pv_terminal}",
                    f"{SHEET_DCF}!B{layout.dcf_enterprise_value}")),
        9: ("WACC", f"={SHEET_WACC}!B{layout.wacc_value}"),
        10: ("WACC less terminal growth",
             formula("difference", f"{SHEET_WACC}!B{layout.wacc_value}",
                     _inp(layout, "terminal_growth"))),
    }
    for row, (label, cell_formula) in rows.items():
        _label(sheet, row, label, bold=row == 4)
        sheet.cell(row, 2, cell_formula)
    sheet.column_dimensions["A"].width = 34


def write_workbook(result: DcfResult, path: str | Path,
                   layout: Layout = LAYOUT, reverse=None, comps=None,
                   target=None, scenarios=None, axes=None) -> Path:
    """Write the valuation as live formulas. Returns the path written.

    `reverse` is an optional `reverse_dcf.ReverseDcf`; when given, its solved
    terminal growth is written to `Inputs` and the Reverse DCF tab rebuilds the
    valuation around it so that Excel checks the solve (4.7).

    `comps` is an optional `CompsExport`; when given, the peer table is written
    to `Inputs` and the Comps tab computes every anchor and the range (4.8).

    `target` is an optional `target.PriceTarget`; when given, the Target tab
    carries the price bridge and the return decomposition (4.9).

    `scenarios` is an optional `scenarios.ScenarioSet` and `axes` a mapping of
    `terminal_growth` and `wacc` to their sensitivity points; each adds its
    tab (4.10).
    """
    if result.conventions != Conventions.SPEC:
        diverging = ", ".join(result.conventions.divergences(Conventions.SPEC))
        raise ExportError(
            f"refusing to export a model whose conventions diverge from the "
            f"framework: {diverging}. reader.py can READ a defective workbook "
            f"-- that is what the audit is for -- but emitting formulas that "
            f"implement the defect would ship it.")

    book = Workbook()
    inputs_sheet = book.active
    inputs_sheet.title = SHEET_INPUTS
    _write_inputs(inputs_sheet, result, layout)
    _write_wacc(book.create_sheet(SHEET_WACC), layout)
    _write_model(book.create_sheet(SHEET_MODEL), result, layout)
    _write_dcf(book.create_sheet(SHEET_DCF), result, layout)
    if comps is not None:
        if len(comps.peers) < 2:
            raise ExportError(
                "a comps tab needs at least two peers, because its deliverable "
                "is the RANGE across anchors and one anchor has no range. "
                "Framework 4.8 requires the full set disclosed.")
        _write_comps_inputs(inputs_sheet, comps, layout)
        _write_comps(book.create_sheet(SHEET_COMPS), layout, comps)
    if scenarios is not None or axes is not None:
        _write_scenario_inputs(inputs_sheet, scenarios, axes, layout)
    if scenarios is not None:
        _write_scenarios(book.create_sheet(SHEET_SCENARIOS), layout, scenarios)
    if axes is not None:
        missing = {"terminal_growth", "wacc"} - set(axes)
        if missing:
            raise ExportError(
                f"a two-way sensitivity table needs both axes; missing "
                f"{sorted(missing)}. 4.10 asks for two-way tables on the "
                f"highest-leverage assumptions, and one axis is not two-way.")
        _write_sensitivity(book.create_sheet(SHEET_SENSITIVITY), layout,
                           len(result.inputs.forecast),
                           len(axes["terminal_growth"]), len(axes["wacc"]))
    if target is not None:
        if target.prior_multiple is None or target.prior_metric is None:
            raise ExportError(
                "the price target has no valuation-date multiple and metric, "
                "so the return decomposition 4.9 requires cannot be built. A "
                "target without it is a number rather than an argument about "
                "where the return comes from.")
        _write_target_inputs(inputs_sheet, target, layout)
        _write_target(book.create_sheet(SHEET_TARGET), layout, target)
    if reverse is not None:
        if reverse.terminal_growth is None:
            raise ExportError(
                "the reverse DCF has no solved terminal growth, so the tab "
                "would have nothing to check. Report it as unsolved instead.")
        _write_reverse_inputs(inputs_sheet, reverse, layout)
        _write_reverse(book.create_sheet(SHEET_REVERSE), layout,
                       len(result.inputs.forecast))
    _write_summary(book.create_sheet(SHEET_SUMMARY), result, layout)

    path = Path(path)
    book.save(path)
    return path
